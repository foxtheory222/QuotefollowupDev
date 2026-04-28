#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook


TOKEN_PATTERN = re.compile(r"\(?\$?-?[0-9,]+(?:\.\d+)?%?\)?|#DIV/0")
WHITESPACE_PATTERN = re.compile(r"\s+")
REFRESH_DATE_PATTERN = re.compile(r"Refresh Date:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2})")
REFRESH_TIME_PATTERN = re.compile(r"Refresh Time:\s*([0-9: ]+[AP]M)\s+(GMT[+-][0-9]{2}:[0-9]{2})")
MONTH_LABEL_PATTERN = re.compile(r"For the Month of ([A-Za-z]+ \d{4})")
METRIC_LINE_PATTERN = re.compile(
    r"^(?P<left>(?:{token})\s+(?:{token})\s+(?:{token}))\s+(?P<label>.+?)\s+(?P<right>(?:{token})\s+(?:{token})\s+(?:{token}))$".format(
        token=<REDACTED>
    )
)

GL060_TOP_LABELS = {
    "Service Center Sales": "sales",
    "Service Center Cost of Sales": "cost_of_sales",
    "Service Center Gross Profit": "gross_profit",
    "Service Center Gross Profit %": "gross_profit_pct",
    "POS Gross Profit %": "pos_gp_pct",
    "Chargeback Gross Profit %": "chargeback_gp_pct",
    "Freight Recovery (Including Delivery Vehicles)": "freight_recovery",
    "Service Center Operating Expenses": "operating_expenses",
    "Service Center Operating Expense %": "operating_expense_pct",
    "Service Center GAP": "gap_pct",
    "OPERATING PROFIT": "operating_profit",
    "NET INCOME AFTER TAXES": "net_income_after_taxes",
}

GL060_DETAIL_LABELS = {
    "Freight Revenue": "freight_revenue",
    "Freight": "freight_subtotal",
    "Fleet Expenses": "fleet_expenses",
    "Compensation & Benefits": "compensation_benefits",
    "Occupancy Expenses": "occupancy_expenses",
    "Shop Expenses": "shop_expenses",
    "Other Operating Expenses": "other_operating_expenses",
    "Administrative Expense": "administrative_expense",
    "Selling Expense": "selling_expense",
    "Office Expense": "office_expense",
    "Other SD&A (Income)/ Expense": "other_sda_income_expense",
}

LOCATION_SUMMARY_FIELDS = [
    "country",
    "area_name",
    "region",
    "region_name",
    "location",
    "location_name",
    "currency",
    "month_end_plan_sales",
    "month_end_plan_gp_pct",
    "mtd_sales",
    "mtd_gp",
    "mtd_gp_pct",
    "last_month_sales",
    "last_month_gp",
    "last_month_gp_pct",
    "ly_month_sales",
    "ly_month_gp",
    "ly_month_gp_pct",
    "sales_change_amount",
    "sales_change_pct",
    "gp_change_pct",
]

ABNORMAL_MARGIN_REVIEW_TYPE_ORDER = {
    "Negative Margin": 10,
    "Low Margin": 20,
    "Large $ Margin": 30,
    "Abnormal High Margin": 40,
}

BRANCHES = {
    "4171": {
        "branch_code": "4171",
        "branch_slug": "4171-calgary",
        "branch_name": "Calgary",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "currency": "CAD",
    },
    "4172": {
        "branch_code": "4172",
        "branch_slug": "4172-lethbridge",
        "branch_name": "Lethbridge",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "currency": "CAD",
    },
    "4173": {
        "branch_code": "4173",
        "branch_slug": "4173-medicine-hat",
        "branch_name": "Medicine Hat",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "currency": "CAD",
    },
}


def normalize_space(value: Any) -> str:
    return WHITESPACE_PATTERN.sub(" ", str(value or "").strip())


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    return text


def as_slug(value: str | None) -> str | None:
    if not value:
        return None
    return re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")


def as_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == "#DIV/0":
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace("$", "").replace(",", "").replace("%", "").replace("(", "").replace(")", "")
    if not cleaned:
        return 0.0
    return -float(cleaned) if negative else float(cleaned)


def as_boolean(value: Any) -> bool:
    return normalize_space(value).lower() in {"1", "true", "yes", "y"}


def as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = as_text(value)
    if text is None:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def as_iso_date(value: Any) -> str | None:
    parsed = as_date(value)
    return parsed.isoformat() if parsed else None


def tokenize_values(text: str) -> list[str]:
    return TOKEN_PATTERN.findall(text or "")


def parse_metric_lines(lines: list[str]) -> dict[str, dict[str, float]]:
    metrics: dict[str, dict[str, float]] = {}
    for raw_line in lines:
        line = normalize_space(raw_line)
        match = METRIC_LINE_PATTERN.match(line)
        if not match:
            continue
        label = normalize_space(match.group("label"))
        left_tokens = tokenize_values(match.group("left"))
        right_tokens = tokenize_values(match.group("right"))
        if len(left_tokens) != 3 or len(right_tokens) != 3:
            continue
        metrics.setdefault(
            label,
            {
                "current_month_actual": as_number(left_tokens[0]),
                "prior_year_month_actual": as_number(left_tokens[1]),
                "plan_month": as_number(left_tokens[2]),
                "current_ytd_actual": as_number(right_tokens[0]),
                "prior_ytd_actual": as_number(right_tokens[1]),
                "plan_ytd": as_number(right_tokens[2]),
            },
        )
    return metrics


def resolve_metric_row(parsed_lines: dict[str, dict[str, float]], label: str) -> dict[str, float] | None:
    if label in parsed_lines:
        return parsed_lines[label]
    for parsed_label, values in parsed_lines.items():
        normalized_label = normalize_space(parsed_label)
        if normalized_label.endswith(label) or f" {label}" in normalized_label:
            return values
    return None


def parse_gl060_refresh(lines: list[str]) -> dict[str, str]:
    refresh_date = None
    refresh_time = None
    refresh_timezone = None
    for line in lines:
        date_match = REFRESH_DATE_PATTERN.search(line)
        if date_match and refresh_date is None:
            refresh_date = datetime.strptime(date_match.group(1), "%m/%d/%y").date()
        time_match = REFRESH_TIME_PATTERN.search(line)
        if time_match and refresh_time is None:
            refresh_time = time_match.group(1)
            refresh_timezone = time_match.group(2)
    if refresh_date is None or refresh_time is None or refresh_timezone is None:
        raise ValueError("GL060 refresh header could not be proven from the PDF.")

    time_value = datetime.strptime(refresh_time, "%I:%M:%S %p").time()
    offset_text = refresh_timezone.replace("GMT", "")
    offset_hours = int(offset_text[0:3])
    offset_minutes = int(offset_text[0] + offset_text[4:6])
    tz = timezone(timedelta(hours=offset_hours, minutes=offset_minutes))
    combined = datetime.combine(refresh_date, time_value, tzinfo=tz)
    return {
        "refresh_date": refresh_date.isoformat(),
        "refresh_time": f"{refresh_time} {refresh_timezone}",
        "refresh_datetime": combined.isoformat(),
    }


def parse_gl060_report(path: Path) -> dict[str, Any]:
    lines: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(text.splitlines())

    refresh = parse_gl060_refresh(lines)
    month_match = next(
        (MONTH_LABEL_PATTERN.search(normalize_space(line)) for line in lines if MONTH_LABEL_PATTERN.search(normalize_space(line))),
        None,
    )
    if month_match is None:
        raise ValueError("GL060 month label could not be proven from the PDF.")
    month_label = month_match.group(1)
    period = datetime.strptime(month_label, "%B %Y")
    parsed_lines = parse_metric_lines(lines)

    top_metrics: dict[str, dict[str, float]] = {}
    for label, key in GL060_TOP_LABELS.items():
        metric_row = resolve_metric_row(parsed_lines, label)
        if metric_row is None:
            raise ValueError(f"GL060 metric '{label}' could not be proven from the PDF.")
        top_metrics[key] = metric_row

    detail_metrics: dict[str, dict[str, float] | None] = {}
    missing_detail_labels: list[str] = []
    for label, key in GL060_DETAIL_LABELS.items():
        metric_row = resolve_metric_row(parsed_lines, label)
        if metric_row is None:
            missing_detail_labels.append(label)
            detail_metrics[key] = None
            continue
        detail_metrics[key] = metric_row

    def detail_amount(key: str) -> float | None:
        row = detail_metrics.get(key)
        if not row:
            return None
        return row["current_month_actual"]

    displayed_freight_recovery = top_metrics["freight_recovery"]["current_month_actual"] / 100.0
    freight_revenue = detail_amount("freight_revenue")
    freight_subtotal = detail_amount("freight_subtotal")
    fleet_expenses = detail_amount("fleet_expenses")
    freight_formula_subtotal_only = (
        freight_revenue / freight_subtotal
        if freight_revenue is not None and freight_subtotal
        else None
    )
    freight_formula_with_fleet = (
        freight_revenue / (freight_subtotal + fleet_expenses)
        if freight_revenue is not None and freight_subtotal is not None and fleet_expenses is not None and (freight_subtotal + fleet_expenses)
        else None
    )

    return {
        "source_report": path.name,
        "month_label": month_label,
        "period_year": period.year,
        "period_month": period.month,
        **refresh,
        "top_metrics": top_metrics,
        "detail_metrics": detail_metrics,
        "waterfall": {
            "sales": top_metrics["sales"]["current_month_actual"],
            "cost_of_sales": top_metrics["cost_of_sales"]["current_month_actual"],
            "gross_profit": top_metrics["gross_profit"]["current_month_actual"],
            "operating_expenses": top_metrics["operating_expenses"]["current_month_actual"],
            "operating_profit": top_metrics["operating_profit"]["current_month_actual"],
        },
        "freight_recovery_discrepancy": {
            "displayed_kpi": displayed_freight_recovery,
            "freight_revenue_divided_by_freight_subtotal": freight_formula_subtotal_only,
            "freight_revenue_divided_by_freight_plus_fleet": freight_formula_with_fleet,
            "notes": "Displayed GL060 KPI reconciles to Freight Revenue / Freight subtotal when those detail rows are present; missing detail labels are recorded instead of aborting fixture generation.",
            "missing_detail_labels": missing_detail_labels,
        },
    }


def parse_sa1300_location_summary(path: Path, branch_code: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Location Summary"]
    rows = list(sheet.iter_rows(values_only=True))
    results: list[dict[str, Any]] = []
    for raw_row in rows[2:]:
        location = as_text(raw_row[4]) if len(raw_row) > 4 else None
        if location is None:
            continue
        if branch_code and location != branch_code:
            continue
        record: dict[str, Any] = {}
        for index, field_name in enumerate(LOCATION_SUMMARY_FIELDS):
            cell = raw_row[index] if index < len(raw_row) else None
            if field_name in {"country", "area_name", "region", "region_name", "location", "location_name", "currency"}:
                record[field_name] = as_text(cell)
            else:
                record[field_name] = as_number(cell) if cell is not None else 0.0
        results.append(record)
    return results


def parse_dual_block_daily_sheet(sheet, block_starts: list[int], block_width: int, includes_location: bool, branch_code: str | None) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    currency_headers = rows[0]
    block_headers = rows[1]
    data_rows = rows[2:]
    results: list[dict[str, Any]] = []

    expected_headers = (
        ["Location", "Billing Day", "Sales", "GP$ (LRMAC)", "GP% (LRMAC)", "On-Time Delivery"]
        if includes_location
        else ["Billing Day", "Sales", "GP$ (LRMAC)", "GP% (LRMAC)", "On-Time Delivery"]
    )

    def header_text(row: tuple[Any, ...], index: int) -> str:
        if index < 0 or index >= len(row):
            return ""
        return normalize_space(row[index])

    def detect_block_starts() -> list[int]:
        starts: list[int] = []
        max_index = len(block_headers)
        for index in range(max_index):
            candidate = [header_text(block_headers, index + offset) for offset in range(len(expected_headers))]
            if candidate == expected_headers:
                starts.append(index)
        return starts or list(block_starts)

    resolved_block_starts = detect_block_starts()
    last_location_by_block: dict[int, str | None] = {start: branch_code for start in resolved_block_starts}

    def parse_block_number(value: Any, row_number: int, field_name: str, start: int) -> float:
        try:
            return as_number(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"{sheet.title}: expected numeric {field_name} in row {row_number}, block {start}, got {value!r}"
            ) from exc

    for row_offset, raw_row in enumerate(data_rows, start=3):
        for start in resolved_block_starts:
            block = raw_row[start : start + block_width]
            if not any(cell is not None and str(cell).strip() for cell in block):
                continue
            currency_header = as_text(currency_headers[start]) or ""
            currency = currency_header.split(" - ")[0].strip()

            if includes_location:
                location = as_text(block[0]) or last_location_by_block.get(start)
                billing_value = block[1]
                sales = parse_block_number(block[2], row_offset, "sales", start)
                gp = parse_block_number(block[3], row_offset, "gp", start)
                gp_pct = parse_block_number(block[4], row_offset, "gp_pct", start)
                on_time_delivery = parse_block_number(block[5], row_offset, "on_time_delivery", start)
            else:
                location = branch_code
                billing_value = block[0]
                sales = parse_block_number(block[1], row_offset, "sales", start)
                gp = parse_block_number(block[2], row_offset, "gp", start)
                gp_pct = parse_block_number(block[3], row_offset, "gp_pct", start)
                on_time_delivery = parse_block_number(block[4], row_offset, "on_time_delivery", start)

            if location is not None:
                last_location_by_block[start] = location
            if branch_code and location and location != branch_code:
                continue

            billing_text = as_text(billing_value)
            is_total_row = billing_text == "Total"
            billing_day = None if is_total_row else as_date(billing_value)
            if not is_total_row and billing_day is None and not sales and not gp and not gp_pct and not on_time_delivery:
                continue

            results.append(
                {
                    "currency": currency,
                    "location": location,
                    "billing_day": billing_day.isoformat() if billing_day else None,
                    "sales": sales,
                    "gp": gp,
                    "gp_pct": gp_pct,
                    "on_time_delivery": on_time_delivery,
                    "is_total_row": is_total_row,
                }
            )
    return results


def parse_sa1300_abnormal_margin(path: Path, branch_code: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Abnormal Margin Review"]
    results: list[dict[str, Any]] = []
    for raw_row in sheet.iter_rows(min_row=3, values_only=True):
        location = as_text(raw_row[3]) if len(raw_row) > 3 else None
        billing_doc = as_text(raw_row[11]) if len(raw_row) > 11 else None
        review_type = as_text(raw_row[6]) if len(raw_row) > 6 else None
        if location is None or billing_doc is None or review_type is None:
            continue
        if branch_code and location != branch_code:
            continue
        results.append(
            {
                "country": as_text(raw_row[0]) if len(raw_row) > 0 else None,
                "area_name": as_text(raw_row[1]) if len(raw_row) > 1 else None,
                "region_name": as_text(raw_row[2]) if len(raw_row) > 2 else None,
                "location": location,
                "location_name": as_text(raw_row[4]) if len(raw_row) > 4 else None,
                "billing_date": as_iso_date(raw_row[5]) if len(raw_row) > 5 else None,
                "review_type": review_type,
                "review_type_slug": as_slug(review_type),
                "currency_type": as_text(raw_row[7]) if len(raw_row) > 7 else None,
                "cssr": as_text(raw_row[8]) if len(raw_row) > 8 else None,
                "cssr_name": as_text(raw_row[9]) if len(raw_row) > 9 else None,
                "customer_name": as_text(raw_row[10]) if len(raw_row) > 10 else None,
                "billing_document_number": billing_doc,
                "billing_document_type": as_text(raw_row[12]) if len(raw_row) > 12 else None,
                "sales": as_number(raw_row[13]) if len(raw_row) > 13 else 0.0,
                "cogs": as_number(raw_row[14]) if len(raw_row) > 14 else 0.0,
                "gp": as_number(raw_row[15]) if len(raw_row) > 15 else 0.0,
                "gp_pct": as_number(raw_row[16]) if len(raw_row) > 16 else 0.0,
            }
        )
    return results


def parse_sa1300_late_orders(path: Path, branch_code: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_name = next(name for name in workbook.sheetnames if name.strip() == "On-Time_ Late Order Review")
    sheet = workbook[sheet_name]
    results: list[dict[str, Any]] = []
    for raw_row in sheet.iter_rows(min_row=3, values_only=True):
        location = as_text(raw_row[4]) if len(raw_row) > 4 else None
        billing_doc = as_text(raw_row[11]) if len(raw_row) > 11 else None
        if location is None or billing_doc is None:
            continue
        if branch_code and location != branch_code:
            continue
        results.append(
            {
                "country_name": as_text(raw_row[0]) if len(raw_row) > 0 else None,
                "area_name": as_text(raw_row[1]) if len(raw_row) > 1 else None,
                "region": as_text(raw_row[2]) if len(raw_row) > 2 else None,
                "region_name": as_text(raw_row[3]) if len(raw_row) > 3 else None,
                "location": location,
                "location_name": as_text(raw_row[5]) if len(raw_row) > 5 else None,
                "cssr": as_text(raw_row[6]) if len(raw_row) > 6 else None,
                "cssr_name": as_text(raw_row[7]) if len(raw_row) > 7 else None,
                "sold_to_customer_name": as_text(raw_row[8]) if len(raw_row) > 8 else None,
                "ship_to_customer_name": as_text(raw_row[9]) if len(raw_row) > 9 else None,
                "billing_date": as_iso_date(raw_row[10]) if len(raw_row) > 10 else None,
                "billing_document_number": billing_doc,
                "material_group": as_text(raw_row[12]) if len(raw_row) > 12 else None,
                "item_category": as_text(raw_row[13]) if len(raw_row) > 13 else None,
                "item_category_description": as_text(raw_row[14]) if len(raw_row) > 14 else None,
                "sales": as_number(raw_row[15]) if len(raw_row) > 15 else 0.0,
                "currency_type": as_text(raw_row[16]) if len(raw_row) > 16 else None,
            }
        )
    return results


def parse_sa1300_report(path: Path, branch_code: str | None) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    daily_sales_location = parse_dual_block_daily_sheet(workbook["Daily Sales- Location"], [0, 7], 6, True, branch_code)
    daily_sales = parse_dual_block_daily_sheet(workbook["Daily Sales"], [0, 6], 5, False, branch_code)
    billing_dates = [as_date(row["billing_day"]) for row in daily_sales_location + daily_sales if row["billing_day"] and not row["is_total_row"]]
    as_of_date = max((value for value in billing_dates if value), default=None)
    return {
        "source_report": path.name,
        "location_summary": parse_sa1300_location_summary(path, branch_code),
        "daily_sales_location": daily_sales_location,
        "daily_sales": daily_sales,
        "abnormal_margin_review": parse_sa1300_abnormal_margin(path, branch_code),
        "late_order_review": parse_sa1300_late_orders(path, branch_code),
        "as_of_date": as_of_date.isoformat() if as_of_date else None,
    }


def parse_sp830_sheet(sheet, gp_column_name: str, branch_code: str | None) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [as_text(cell) or "" for cell in rows[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    line_rows: list[dict[str, Any]] = []
    for raw_row in rows[2:]:
        quote_number = as_text(raw_row[header_index["Quote Number"]]) if "Quote Number" in header_index else None
        if quote_number is None:
            continue
        location = as_text(raw_row[header_index.get("Location", header_index.get("Location ", -1))])
        if branch_code and location != branch_code:
            continue
        line_rows.append(
            {
                "created_on": as_iso_date(raw_row[header_index["Created On"]]) if "Created On" in header_index else None,
                "location": location,
                "quote_number": quote_number,
                "item_number": as_text(raw_row[header_index.get("Item Number", header_index.get("Item Number ", -1))]),
                "sold_to_party": as_text(raw_row[header_index["Sold-To Party"]]) if "Sold-To Party" in header_index else None,
                "sold_to_party_name": as_text(raw_row[header_index.get("Sold To Party Name", header_index.get("Sold To Party Name ", -1))]),
                "cssr": as_text(raw_row[header_index["CSSR"]]) if "CSSR" in header_index else None,
                "cssr_name": as_text(raw_row[header_index["CSSR Name"]]) if "CSSR Name" in header_index else None,
                "am_number": as_text(raw_row[header_index["AM Number"]]) if "AM Number" in header_index else None,
                "am_name": as_text(raw_row[header_index["AM Name"]]) if "AM Name" in header_index else None,
                "material_group": as_text(raw_row[header_index["Material Group"]]) if "Material Group" in header_index else None,
                "manufacturer_part_description": as_text(raw_row[header_index["Manufacturer Part Description"]]) if "Manufacturer Part Description" in header_index else None,
                "value": as_number(raw_row[header_index["Value"]]) if "Value" in header_index else 0.0,
                "gp_percent": as_number(raw_row[header_index[gp_column_name]]) if gp_column_name in header_index else 0.0,
                "non_converted_quote": as_number(raw_row[header_index["$ NonConv Quote"]]) if "$ NonConv Quote" in header_index else 0.0,
                "converted_quote": as_number(raw_row[header_index["$ Converted Quote"]]) if "$ Converted Quote" in header_index else 0.0,
                "rejection_reason": as_text(raw_row[header_index["Rejection Reason"]]) if "Rejection Reason" in header_index else None,
                "follow_up_owner": as_text(raw_row[header_index.get("Follow up Owner", header_index.get("Follow up Owner ", -1))]),
                "status": as_text(raw_row[header_index["Status (Won/Loss)"]]) if "Status (Won/Loss)" in header_index else None,
                "additional_information": as_text(raw_row[header_index["Additional Information"]]) if "Additional Information" in header_index else None,
            }
        )
    created_dates = [as_date(row["created_on"]) for row in line_rows if row["created_on"]]
    return {
        "line_rows": line_rows,
        "header_row_index": 2,
        "as_of_date": max(created_dates).isoformat() if created_dates else None,
    }


def parse_sp830_report(path: Path, branch_code: str | None) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    return {
        "source_report": path.name,
        "daily": parse_sp830_sheet(workbook["Daily"], "GP% Rep", branch_code),
        "monthly": parse_sp830_sheet(workbook["Monthly"], "GP% Rep_CM", branch_code),
    }


def parse_zbo_report(path: Path, branch_code: str | None) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["ZBO"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [as_text(cell) or "" for cell in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    line_rows: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        soff = as_text(raw_row[header_index["Soff."]]) if "Soff." in header_index else None
        sales_doc = as_text(raw_row[header_index["Sales Doc #"]]) if "Sales Doc #" in header_index else None
        line_number = as_text(raw_row[header_index["Line"]]) if "Line" in header_index else None
        if soff is None or sales_doc is None or line_number is None:
            continue
        if branch_code and soff != branch_code:
            continue
        line_rows.append(
            {
                "branch_code": soff,
                "on_time_date": as_iso_date(raw_row[header_index["On-Time Date"]]) if "On-Time Date" in header_index else None,
                "sales_doc_number": sales_doc,
                "line_number": line_number,
                "sales_doc_type": as_text(raw_row[header_index["Sales Doc Type"]]) if "Sales Doc Type" in header_index else None,
                "customer_po_number": as_text(raw_row[header_index["Customer PO Number"]]) if "Customer PO Number" in header_index else None,
                "material_group": as_text(raw_row[header_index["Mat Grp"]]) if "Mat Grp" in header_index else None,
                "material": as_text(raw_row[header_index["Material"]]) if "Material" in header_index else None,
                "material_description": as_text(raw_row[header_index["Material Description"]]) if "Material Description" in header_index else None,
                "account_manager": as_text(raw_row[header_index["Acct Mgr"]]) if "Acct Mgr" in header_index else None,
                "account_manager_name": as_text(raw_row[header_index["Acct Mgr Name"]]) if "Acct Mgr Name" in header_index else None,
                "sold_to": as_text(raw_row[header_index["Sold-To"]]) if "Sold-To" in header_index else None,
                "sold_to_name": as_text(raw_row[header_index["Sold To Name"]]) if "Sold To Name" in header_index else None,
                "ship_to": as_text(raw_row[header_index["Ship To"]]) if "Ship To" in header_index else None,
                "ship_to_name": as_text(raw_row[header_index["Ship To Name"]]) if "Ship To Name" in header_index else None,
                "unshipped_net_value": as_number(raw_row[header_index["Unshipped Net Value"]]) if "Unshipped Net Value" in header_index else 0.0,
                "cumulative_order_quantity": as_number(raw_row[header_index["Cumulative Order Quantity "]]) if "Cumulative Order Quantity " in header_index else 0.0,
                "shipped_qty": as_number(raw_row[header_index["Shipped Qty"]]) if "Shipped Qty" in header_index else 0.0,
                "unshipped_quantity": as_number(raw_row[header_index["Unshipped Quantity"]]) if "Unshipped Quantity" in header_index else 0.0,
                "qty_billed": as_number(raw_row[header_index["Qty Billed"]]) if "Qty Billed" in header_index else 0.0,
                "qty_on_del_not_pgid": as_number(raw_row[header_index["Qty on Del Not PGI'd"]]) if "Qty on Del Not PGI'd" in header_index else 0.0,
                "qty_not_on_del": as_number(raw_row[header_index["Qty Not On Del"]]) if "Qty Not On Del" in header_index else 0.0,
                "uom": as_text(raw_row[header_index["UOM"]]) if "UOM" in header_index else None,
                "net_price": as_number(raw_row[header_index["Net Price"]]) if "Net Price" in header_index else 0.0,
                "ship_cond_desc": as_text(raw_row[header_index["Ship Cond Desc"]]) if "Ship Cond Desc" in header_index else None,
                "line_item_created_on_date": as_iso_date(raw_row[header_index["Line Item Created On Date"]]) if "Line Item Created On Date" in header_index else None,
                "first_date": as_iso_date(raw_row[header_index["First date"]]) if "First date" in header_index else None,
                "del_block_desc": as_text(raw_row[header_index["Del Block Desc"]]) if "Del Block Desc" in header_index else None,
                "bill_block_desc": as_text(raw_row[header_index["Bill Block Desc"]]) if "Bill Block Desc" in header_index else None,
                "item_category": as_text(raw_row[header_index["Item Category"]]) if "Item Category" in header_index else None,
                "vendor_po": as_text(raw_row[header_index["Vendor PO"]]) if "Vendor PO" in header_index else None,
                "cssr_name": as_text(raw_row[header_index["CSSR Name"]]) if "CSSR Name" in header_index else None,
                "created_by": as_text(raw_row[header_index["Created By"]]) if "Created By" in header_index else None,
                "plant": as_text(raw_row[header_index["Plant"]]) if "Plant" in header_index else None,
                "user_status_description": as_text(raw_row[header_index["User Status Description"]]) if "User Status Description" in header_index else None,
                "reason_for_rejection": as_text(raw_row[header_index["Reason for Rejection"]]) if "Reason for Rejection" in header_index else None,
            }
        )
    created_dates = [as_date(row["line_item_created_on_date"]) for row in line_rows if row["line_item_created_on_date"]]
    return {"source_report": path.name, "line_rows": line_rows, "as_of_date": max(created_dates).isoformat() if created_dates else None}


def parse_freight_report(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["CustomerBill - Filter Report"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [as_text(cell) or "" for cell in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    shipment_rows: list[dict[str, Any]] = []

    for raw_row in rows[1:]:
        invoice_number = as_text(raw_row[header_index["Invoice Number"]]) if "Invoice Number" in header_index else None
        if invoice_number is None:
            continue
        shipment_rows.append(
            {
                "invoice_number": invoice_number,
                "invoice_carrier": as_text(raw_row[header_index["Invoice Carrier"]]) if "Invoice Carrier" in header_index else None,
                "carrier_mode": as_text(raw_row[header_index["Carrier Mode"]]) if "Carrier Mode" in header_index else None,
                "pro": as_text(raw_row[header_index["PRO"]]) if "PRO" in header_index else None,
                "target_ship_early": as_iso_date(raw_row[header_index["Target Ship (Early)"]]) if "Target Ship (Early)" in header_index else None,
                "actual_ship": as_iso_date(raw_row[header_index["Actual Ship"]]) if "Actual Ship" in header_index else None,
                "target_delivery_early": as_iso_date(raw_row[header_index["Target Delivery (Early)"]]) if "Target Delivery (Early)" in header_index else None,
                "actual_delivery": as_iso_date(raw_row[header_index["Actual Delivery"]]) if "Actual Delivery" in header_index else None,
                "invoice_weight": as_number(raw_row[header_index["Invoice Weight"]]) if "Invoice Weight" in header_index else 0.0,
                "weight": as_number(raw_row[header_index["Weight"]]) if "Weight" in header_index else 0.0,
                "freight_cost_only": as_number(raw_row[header_index["Freight Cost Only"]]) if "Freight Cost Only" in header_index else 0.0,
                "line_haul": as_number(raw_row[header_index["Line Haul"]]) if "Line Haul" in header_index else 0.0,
                "fuel": as_number(raw_row[header_index["Fuel"]]) if "Fuel" in header_index else 0.0,
                "accessorial_cts": as_number(raw_row[header_index["Accessorial - CTS"]]) if "Accessorial - CTS" in header_index else 0.0,
                "accessorial_msc": as_number(raw_row[header_index["Accessorial - MSC"]]) if "Accessorial - MSC" in header_index else 0.0,
                "acc_1_amount": as_number(raw_row[header_index["ACC 1 Amount"]]) if "ACC 1 Amount" in header_index else 0.0,
                "acc_2_amount": as_number(raw_row[header_index["ACC 2 Amount"]]) if "ACC 2 Amount" in header_index else 0.0,
                "acc_3_amount": as_number(raw_row[header_index["ACC 3 Amount"]]) if "ACC 3 Amount" in header_index else 0.0,
                "acc_4_amount": as_number(raw_row[header_index["ACC 4 Amount"]]) if "ACC 4 Amount" in header_index else 0.0,
                "least_cost_carrier": as_text(raw_row[header_index["Least Cost Carrier"]]) if "Least Cost Carrier" in header_index else None,
                "least_cost_normalized_total": as_number(raw_row[header_index["Least Cost Normalized Total"]]) if "Least Cost Normalized Total" in header_index else 0.0,
                "least_cost_service_days": as_number(raw_row[header_index["Least Cost Service Days"]]) if "Least Cost Service Days" in header_index else 0.0,
                "actual_transit_days": as_number(raw_row[header_index["Actual Transit Days"]]) if "Actual Transit Days" in header_index else 0.0,
                "preferred_carrier_used": as_boolean(raw_row[header_index["Preferred Carrier Used"]]) if "Preferred Carrier Used" in header_index else False,
                "normalized_unrealized_savings": as_number(raw_row[header_index["Normalized Unrealized Savings"]]) if "Normalized Unrealized Savings" in header_index else 0.0,
            }
        )

    freshness_dates = [
        as_date(row["actual_delivery"]) or as_date(row["actual_ship"]) or as_date(row["target_delivery_early"])
        for row in shipment_rows
    ]
    freshness_values = [value for value in freshness_dates if value is not None]
    return {
        "source_report": path.name,
        "shipment_rows": shipment_rows,
        "as_of_date": max(freshness_values).isoformat() if freshness_values else None,
        "freshness_rule": "max(actual_delivery, else actual_ship, else target_delivery_early)",
    }


def missing_freight_report() -> dict[str, Any]:
    return {
        "source_report": None,
        "shipment_rows": [],
        "as_of_date": None,
        "freshness_rule": "awaiting example freight workbook",
        "status": "awaiting_feed",
    }


def safe_ratio(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    return numerator / denominator


def summarize_ops_daily(sa1300_report: dict[str, Any]) -> dict[str, Any]:
    location_rows = [row for row in sa1300_report["location_summary"] if row["currency"] == "CAD"]
    summary_row = location_rows[0] if location_rows else (sa1300_report["location_summary"][0] if sa1300_report["location_summary"] else None)
    daily_rows = [
        row
        for row in sa1300_report["daily_sales_location"]
        if row["currency"] == "CAD" and not row["is_total_row"]
    ]
    total_row = next(
        (row for row in sa1300_report["daily_sales_location"] if row["currency"] == "CAD" and row["is_total_row"]),
        None,
    )
    trend_rows = sorted(daily_rows, key=lambda row: row["billing_day"] or "")
    return {
        "as_of_date": sa1300_report["as_of_date"],
        "location_summary": summary_row,
        "daily_trend_rows": trend_rows,
        "daily_total_row": total_row,
    }


def summarize_quote_lines(line_rows: list[dict[str, Any]], as_of_date: date) -> dict[str, Any]:
    distinct_quotes = {row["quote_number"] for row in line_rows if row["quote_number"]}
    quote_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"converted": 0.0, "value": 0.0})
    aging_buckets = {"0-7": 0.0, "8-14": 0.0, "15-30": 0.0, "31+": 0.0}
    cssr_ranking: dict[str, float] = defaultdict(float)
    am_ranking: dict[str, float] = defaultdict(float)
    rejection_reasons: dict[str, int] = defaultdict(int)
    open_quote_value = 0.0
    converted_quote_value = 0.0
    weighted_gp_numerator = 0.0
    weighted_gp_denominator = 0.0

    for row in line_rows:
        quote_number = row["quote_number"]
        value = row["value"]
        open_quote_value += row["non_converted_quote"]
        converted_quote_value += row["converted_quote"]
        quote_totals[quote_number]["converted"] += row["converted_quote"]
        quote_totals[quote_number]["value"] += value
        weighted_gp_numerator += value * row["gp_percent"]
        weighted_gp_denominator += value

        created_on = as_date(row["created_on"])
        if created_on is not None:
            age_days = max((as_of_date - created_on).days, 0)
            if age_days <= 7:
                aging_buckets["0-7"] += value
            elif age_days <= 14:
                aging_buckets["8-14"] += value
            elif age_days <= 30:
                aging_buckets["15-30"] += value
            else:
                aging_buckets["31+"] += value

        if row["cssr_name"]:
            cssr_ranking[row["cssr_name"]] += value
        if row["am_name"]:
            am_ranking[row["am_name"]] += value
        if row["rejection_reason"]:
            rejection_reasons[row["rejection_reason"]] += 1

    converted_quote_count = sum(1 for value in quote_totals.values() if value["converted"] > 0)
    return {
        "line_rows": len(line_rows),
        "distinct_quotes": len(distinct_quotes),
        "open_quote_value": round(open_quote_value, 2),
        "converted_quote_value": round(converted_quote_value, 2),
        "conversion_pct_by_value": safe_ratio(converted_quote_value, converted_quote_value + open_quote_value),
        "converted_quote_count": converted_quote_count,
        "conversion_pct_by_count": safe_ratio(converted_quote_count, len(distinct_quotes)),
        "weighted_quote_gp_pct": safe_ratio(weighted_gp_numerator, weighted_gp_denominator),
        "aging_buckets": {key: round(value, 2) for key, value in aging_buckets.items()},
        "cssr_ranking": dict(sorted(cssr_ranking.items(), key=lambda item: item[1], reverse=True)),
        "am_ranking": dict(sorted(am_ranking.items(), key=lambda item: item[1], reverse=True)),
        "rejection_reasons": dict(sorted(rejection_reasons.items(), key=lambda item: item[1], reverse=True)),
    }


def summarize_zbo_lines(line_rows: list[dict[str, Any]], as_of_date: date) -> dict[str, Any]:
    open_rows = [
        row
        for row in line_rows
        if row["unshipped_quantity"] > 0 and not normalize_space(row["reason_for_rejection"] or "")
    ]
    overdue_rows = []
    blocked_orders: set[str] = set()
    pending_approval_value = 0.0
    must_add_freight_value = 0.0
    ship_condition_mix: dict[str, float] = defaultdict(float)

    for row in open_rows:
        on_time_date = as_date(row["on_time_date"])
        if on_time_date is not None and on_time_date < as_of_date:
            overdue_rows.append(row)
        if row["del_block_desc"] or row["bill_block_desc"]:
            blocked_orders.add(row["sales_doc_number"])
        if normalize_space(row["del_block_desc"] or "").upper() == "PENDING APPROVAL":
            pending_approval_value += row["unshipped_net_value"]
        bill_block_text = normalize_space(row["bill_block_desc"] or "").upper()
        del_block_text = normalize_space(row["del_block_desc"] or "").upper()
        if del_block_text == "MUST ADD FREIGHT" or "FREIGHT" in bill_block_text:
            must_add_freight_value += row["unshipped_net_value"]
        ship_condition_mix[row["ship_cond_desc"] or "Unspecified"] += row["unshipped_net_value"]

    return {
        "distinct_open_orders": len({row["sales_doc_number"] for row in open_rows}),
        "open_backlog_value": round(sum(row["unshipped_net_value"] for row in open_rows), 2),
        "overdue_backlog_value": round(sum(row["unshipped_net_value"] for row in overdue_rows), 2),
        "blocked_orders_distinct_count": len(blocked_orders),
        "pending_approval_value": round(pending_approval_value, 2),
        "must_add_freight_value": round(must_add_freight_value, 2),
        "ship_condition_mix": {key: round(value, 2) for key, value in sorted(ship_condition_mix.items())},
        "blocked_rows": [row for row in open_rows if row["sales_doc_number"] in blocked_orders],
        "as_of_date": max((row["line_item_created_on_date"] for row in open_rows if row["line_item_created_on_date"]), default=None),
    }


def summarize_margin_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts: dict[str, int] = defaultdict(int)
    sales_by_type: dict[str, float] = defaultdict(float)
    gp_by_type: dict[str, float] = defaultdict(float)
    for row in rows:
        review_type = row["review_type"] or "Unknown"
        counts[review_type] += 1
        sales_by_type[review_type] += row["sales"]
        gp_by_type[review_type] += row["gp"]
    return {
        "total_rows": len(rows),
        "counts_by_type": dict(sorted(counts.items(), key=lambda item: ABNORMAL_MARGIN_REVIEW_TYPE_ORDER.get(item[0], 999))),
        "sales_by_type": {key: round(value, 2) for key, value in sorted(sales_by_type.items(), key=lambda item: ABNORMAL_MARGIN_REVIEW_TYPE_ORDER.get(item[0], 999))},
        "gp_by_type": {key: round(value, 2) for key, value in sorted(gp_by_type.items(), key=lambda item: ABNORMAL_MARGIN_REVIEW_TYPE_ORDER.get(item[0], 999))},
        "total_sales": round(sum(row["sales"] for row in rows), 2),
        "total_gp": round(sum(row["gp"] for row in rows), 2),
    }


def summarize_late_orders(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "total_rows": len(rows),
        "total_sales": round(sum(row["sales"] for row in rows), 2),
    }


def summarize_freight_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    accessorial_amount = 0.0
    freight_spend = 0.0
    preferred_used = 0
    weight_total = 0.0
    transit_variances: list[float] = []
    carrier_spend: dict[str, float] = defaultdict(float)

    for row in rows:
        freight_spend += row["freight_cost_only"]
        accessorial_amount += (
            row["accessorial_cts"]
            + row["accessorial_msc"]
            + row["acc_1_amount"]
            + row["acc_2_amount"]
            + row["acc_3_amount"]
            + row["acc_4_amount"]
        )
        preferred_used += 1 if row["preferred_carrier_used"] else 0
        weight_total += row["weight"] if row["weight"] > 0 else row["invoice_weight"]
        if row["actual_transit_days"] or row["least_cost_service_days"]:
            transit_variances.append(row["actual_transit_days"] - row["least_cost_service_days"])
        carrier_spend[row["invoice_carrier"] or "Unknown"] += row["freight_cost_only"]

    shipment_count = len(rows)
    return {
        "shipment_row_count": shipment_count,
        "distinct_invoice_count": len({row["invoice_number"] for row in rows}),
        "freight_spend": round(freight_spend, 2),
        "average_cost_per_shipment": round(safe_ratio(freight_spend, shipment_count), 2),
        "fuel_pct": safe_ratio(sum(row["fuel"] for row in rows), freight_spend),
        "accessorial_amount": round(accessorial_amount, 2),
        "accessorial_pct": safe_ratio(accessorial_amount, freight_spend),
        "preferred_carrier_usage_pct": safe_ratio(preferred_used, shipment_count),
        "unrealized_savings": round(sum(row["normalized_unrealized_savings"] for row in rows), 2),
        "cost_per_lb": safe_ratio(freight_spend, weight_total),
        "transit_variance_days_avg": safe_ratio(sum(transit_variances), len(transit_variances)),
        "carrier_spend": {key: round(value, 2) for key, value in sorted(carrier_spend.items(), key=lambda item: item[0])},
    }


def build_branch_analytics_payload(example_root: Path, branch_code: str, zbo_as_of_date: date | None = None) -> dict[str, Any]:
    branch = BRANCHES[branch_code]
    branch_root = example_root / branch_code
    gl060_path = next(branch_root.glob("GL060 Report*.pdf"))
    sa1300_path = next(branch_root.glob("SA1300*.xlsx"))
    sp830_path = next(branch_root.glob("*Quote Follow Up Report*.xlsx"))
    zbo_path = next(branch_root.glob("*ZBO*.xlsx"))
    freight_path = next((branch_root / "Other").glob("*Invoice Report*.xlsx"), None)

    gl060 = parse_gl060_report(gl060_path)
    sa1300 = parse_sa1300_report(sa1300_path, branch_code)
    sp830 = parse_sp830_report(sp830_path, branch_code)
    zbo = parse_zbo_report(zbo_path, branch_code)
    freight = parse_freight_report(freight_path) if freight_path else missing_freight_report()

    sp830_month_as_of = as_date(sp830["monthly"]["as_of_date"]) or as_date(sa1300["as_of_date"]) or date.today()
    zbo_overdue_as_of = zbo_as_of_date or date.today()

    return {
        "branch": branch,
        "gl060": gl060,
        "sa1300": sa1300,
        "sp830": sp830,
        "zbo": zbo,
        "freight": freight,
        "summaries": {
            "ops_daily": summarize_ops_daily(sa1300),
            "sp830_monthly": summarize_quote_lines(sp830["monthly"]["line_rows"], sp830_month_as_of),
            "sp830_daily": summarize_quote_lines(sp830["daily"]["line_rows"], sp830_month_as_of),
            "zbo": summarize_zbo_lines(zbo["line_rows"], zbo_overdue_as_of),
            "abnormal_margin": summarize_margin_rows(sa1300["abnormal_margin_review"]),
            "late_orders": summarize_late_orders(sa1300["late_order_review"]),
            "freight": summarize_freight_rows(freight["shipment_rows"]),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--example-root", default="example")
    parser.add_argument("--branch-code")
    parser.add_argument("--as-of-date")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    example_root = Path(args.example_root)
    branch_codes = [args.branch_code] if args.branch_code else sorted(BRANCHES)
    zbo_as_of_date = as_date(args.as_of_date) if args.as_of_date else None
    payload = {
        "generated_on": datetime.now().isoformat(),
        "branches": [build_branch_analytics_payload(example_root, branch_code, zbo_as_of_date=zbo_as_of_date) for branch_code in branch_codes],
    }

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()

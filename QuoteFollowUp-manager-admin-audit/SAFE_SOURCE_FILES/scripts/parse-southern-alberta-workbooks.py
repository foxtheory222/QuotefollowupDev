#!/usr/bin/env python3

import argparse
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


BRANCHES = {
    "4171": {
        "branch_code": "4171",
        "branch_slug": "4171-calgary",
        "branch_name": "Calgary",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "mailbox_address": "<EMAIL>",
        "sort_order": 10,
    },
    "4172": {
        "branch_code": "4172",
        "branch_slug": "4172-lethbridge",
        "branch_name": "Lethbridge",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "mailbox_address": "<EMAIL>",
        "sort_order": 20,
    },
    "4173": {
        "branch_code": "4173",
        "branch_slug": "4173-medicine-hat",
        "branch_name": "Medicine Hat",
        "region_slug": "southern-alberta",
        "region_name": "Southern Alberta",
        "mailbox_address": "<EMAIL>",
        "sort_order": 30,
    },
}


def as_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("#"):
        return None
    return text


def as_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.startswith("#"):
        return 0.0
    text = text.replace("$", "").replace(",", "").replace("%", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_int(value):
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = as_text(value)
    if text is None:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def as_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = as_text(value)
    if text is None:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def iso_date(value):
    if value is None:
        return None
    return value.isoformat()


def compact_json(value):
    return json.dumps(value, separators=(",", ":"))


def find_sheet_name(workbook, target_name):
    for name in workbook.sheetnames:
        if name.strip() == target_name:
            return name
    raise KeyError(f"Worksheet not found: {target_name}")


def workbook_json_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip()


def currency_label(value):
    text = as_text(value)
    if text is None:
        return None
    lowered = text.lower()
    if "currency" not in lowered:
        return None
    if "cad" in lowered:
        return "CAD"
    if "usd" in lowered:
        return "USD"
    return None


def parse_ops_daily_block(rows, branch_code, start_index, currency):
    block_headers = []
    header_row = rows[1] if len(rows) > 1 else []
    for offset in range(6):
        header_index = start_index + offset
        header_value = header_row[header_index] if header_index < len(header_row) else None
        block_headers.append(as_text(header_value))

    if block_headers[:2] != ["Location", "Billing Day"]:
        return []

    records = []
    for row in rows[2:]:
        values = []
        for offset in range(6):
            value_index = start_index + offset
            values.append(row[value_index] if value_index < len(row) else None)

        if all(value is None or as_text(value) is None for value in values):
            continue

        record = {header: workbook_json_value(value) for header, value in zip(block_headers, values) if header}
        billing_day = record.get("Billing Day")
        if billing_day is None:
            continue

        location = as_text(record.get("Location"))
        if location != branch_code and billing_day != "Total":
            continue

        record["Currency"] = currency
        records.append(record)

    return records


def parse_ops_daily(path, branch):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[find_sheet_name(workbook, "Daily Sales- Location")]
    rows = list(sheet.iter_rows(values_only=True))
    first_row = rows[0] if rows else []

    currency_blocks = []
    for index, value in enumerate(first_row):
        currency = currency_label(value)
        if currency is None:
            continue
        currency_blocks.append((currency, index))

    currency_rows = defaultdict(list)
    for currency, start_index in currency_blocks:
        for record in parse_ops_daily_block(rows, branch["branch_code"], start_index, currency):
            currency_rows[currency].append(record)

    all_records = []
    latest_billing_day = None
    snapshot_date = datetime.fromtimestamp(path.stat().st_mtime).date()
    captured_on = datetime.fromtimestamp(path.stat().st_mtime).isoformat()

    for currency in ("CAD", "USD"):
        rows_for_currency = currency_rows.get(currency, [])
        for row in rows_for_currency:
            billing_label = row.get("Billing Day")
            is_total = billing_label == "Total"
            billing_date = None if is_total else as_date(billing_label)
            if billing_date is not None and (latest_billing_day is None or billing_date > latest_billing_day):
                latest_billing_day = billing_date

            billing_key = "total" if is_total else billing_date.isoformat()
            all_records.append(
                {
                    "qfu_name": f"{branch['branch_code']} Ops Daily {currency} {'Total' if is_total else billing_key}",
                    "qfu_sourceid": f"{branch['branch_code']}|SA1300-OPSDAILY|{snapshot_date.isoformat()}|{currency}|{billing_key}",
                    "qfu_branchcode": branch["branch_code"],
                    "qfu_branchslug": branch["branch_slug"],
                    "qfu_regionslug": branch["region_slug"],
                    "qfu_sourcefamily": "SA1300-OPSDAILY",
                    "qfu_sourcefile": path.name,
                    "qfu_sourceworksheet": "Daily Sales- Location",
                    "qfu_snapshotdate": snapshot_date.isoformat(),
                    "qfu_billingday": None if is_total else billing_key,
                    "qfu_billinglabel": "Total" if is_total else billing_key,
                    "qfu_istotalrow": is_total,
                    "qfu_currencytype": currency,
                    "qfu_sales": round(as_float(row.get("Sales")), 2),
                    "qfu_gp": round(as_float(row.get("GP$ (LRMAC)")), 2),
                    "qfu_gppct": round(as_float(row.get("GP% (LRMAC)")) * 100, 4),
                    "qfu_ontimedelivery": round(as_float(row.get("On-Time Delivery")) * 100, 4),
                    "qfu_sortorder": 999 if is_total else billing_date.day,
                }
            )

    return {
        "file_name": path.name,
        "captured_on": captured_on,
        "snapshot_date": snapshot_date.isoformat(),
        "latest_billing_day": iso_date(latest_billing_day),
        "cad_rows": currency_rows.get("CAD", []),
        "usd_rows": currency_rows.get("USD", []),
        "records": sorted(all_records, key=lambda item: (item["qfu_currencytype"], item["qfu_sortorder"])),
    }


def normalize_status(status_text, converted_quote, rejection_reason):
    text = (status_text or "").strip().lower()
    if "won" in text or converted_quote > 0:
        return 2
    if "lost" in text or rejection_reason:
        return 3
    return 1


def first_non_empty(*values):
    for value in values:
        text = as_text(value)
        if text is not None:
            return text
    return None


def row_value(row, headers, *names):
    for name in names:
        if name in headers:
            index = headers[name]
            if index >= 0 and index < len(row):
                return row[index]
    return None


def read_rows(sheet, header_row_index):
    raw_rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell).strip() if cell is not None else "" for cell in raw_rows[header_row_index]]
    keys = {name: idx for idx, name in enumerate(header)}
    return keys, raw_rows[header_row_index + 1 :]


def parse_quote_file(path, branch):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Daily"]
    headers, rows = read_rows(sheet, 1)
    aggregated = {}

    for row in rows:
        quote_number = as_text(row[headers.get("Quote Number", -1)]) if "Quote Number" in headers else None
        if quote_number is None:
            continue

        created_on = as_date(row[headers.get("Created On", -1)]) if "Created On" in headers else None
        value = as_float(row[headers.get("Value", -1)]) if "Value" in headers else 0.0
        converted_quote = as_float(row[headers.get("$ Converted Quote", -1)]) if "$ Converted Quote" in headers else 0.0
        rejection_reason = as_text(row[headers.get("Rejection Reason", -1)]) if "Rejection Reason" in headers else None
        status_text = as_text(row[headers.get("Status (Won/Loss)", -1)]) if "Status (Won/Loss)" in headers else None
        status_code = normalize_status(status_text, converted_quote, rejection_reason)
        owner = first_non_empty(
            row[headers.get("Follow up Owner", -1)] if "Follow up Owner" in headers else None,
            row[headers.get("Follow up Owner ", -1)] if "Follow up Owner " in headers else None,
            row[headers.get("AM Name", -1)] if "AM Name" in headers else None,
            row[headers.get("CSSR Name", -1)] if "CSSR Name" in headers else None,
        )

        existing = aggregated.get(quote_number)
        if existing is None:
            aggregated[quote_number] = {
                "qfu_name": f"{branch['branch_code']} {quote_number}",
                "qfu_sourceid": f"{branch['branch_code']}|SP830CA|{quote_number}",
                "qfu_quotenumber": quote_number,
                "qfu_customername": first_non_empty(
                    row[headers.get("Sold To Party Name", -1)] if "Sold To Party Name" in headers else None,
                    row[headers.get("Sold To Party Name ", -1)] if "Sold To Party Name " in headers else None,
                ),
                "qfu_amount": value,
                "qfu_assignedto": owner,
                "qfu_cssrname": first_non_empty(row[headers.get("CSSR Name", -1)] if "CSSR Name" in headers else None),
                "qfu_nextfollowup": None,
                "qfu_overduesince": iso_date(created_on),
                "qfu_lasttouchedon": None,
                "qfu_lastfollowupupdatedon": None,
                "qfu_priorityscore": 0,
                "qfu_actionstate": 0,
                "qfu_status": status_code,
                "qfu_sourcedate": iso_date(created_on),
                "qfu_sourceupdatedon": iso_date(created_on),
                "qfu_branchcode": branch["branch_code"],
                "qfu_branchslug": branch["branch_slug"],
                "qfu_regionslug": branch["region_slug"],
                "qfu_sourcefamily": "SP830CA",
                "qfu_sourcefile": path.name,
                "qfu_sourceworksheet": "Daily",
                "qfu_source_row_count": 1,
            }
            continue

        existing["qfu_amount"] = round(existing["qfu_amount"] + value, 2)
        existing["qfu_source_row_count"] += 1
        if existing["qfu_customername"] is None:
            existing["qfu_customername"] = first_non_empty(
                row[headers.get("Sold To Party Name", -1)] if "Sold To Party Name" in headers else None,
                row[headers.get("Sold To Party Name ", -1)] if "Sold To Party Name " in headers else None,
            )
        if existing["qfu_assignedto"] is None:
            existing["qfu_assignedto"] = owner
        if existing["qfu_cssrname"] is None:
            existing["qfu_cssrname"] = first_non_empty(row[headers.get("CSSR Name", -1)] if "CSSR Name" in headers else None)

        existing_status = existing["qfu_status"]
        if status_code == 2 or (status_code == 3 and existing_status == 1):
            existing["qfu_status"] = status_code

        current_source_date = as_date(existing["qfu_sourcedate"])
        if created_on is not None:
            if current_source_date is None or created_on < current_source_date:
                existing["qfu_sourcedate"] = iso_date(created_on)
                existing["qfu_overduesince"] = iso_date(created_on)
            current_updated = as_date(existing["qfu_sourceupdatedon"])
            if current_updated is None or created_on > current_updated:
                existing["qfu_sourceupdatedon"] = iso_date(created_on)

    for record in aggregated.values():
        source_date = as_date(record["qfu_sourcedate"])
        if source_date is not None:
            today = date.today()
            age_days = max((today - source_date).days, 0)
            record["qfu_priorityscore"] = int(round(record["qfu_amount"])) + age_days

    return sorted(aggregated.values(), key=lambda item: (item["qfu_quotenumber"] or ""))


def parse_quote_lines(path, branch):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Daily"]
    headers, rows = read_rows(sheet, 1)
    captured_on = datetime.fromtimestamp(path.stat().st_mtime).isoformat()
    line_records = []

    for row in rows:
        quote_number = as_text(row_value(row, headers, "Quote Number"))
        if quote_number is None:
            continue

        location = as_text(row_value(row, headers, "Location", "Location "))
        if location and location != branch["branch_code"]:
            continue

        line_number = as_text(row_value(row, headers, "Item Number", "Item Number ")) or "0"
        created_on = as_date(row_value(row, headers, "Created On"))
        converted_quote = round(as_float(row_value(row, headers, "$ Converted Quote")), 2)
        total_value = round(as_float(row_value(row, headers, "Value")), 2)
        status_text = as_text(row_value(row, headers, "Status (Won/Loss)"))
        rejection_reason = as_text(row_value(row, headers, "Rejection Reason"))
        status_code = normalize_status(status_text, converted_quote, rejection_reason)

        line_records.append(
            {
                "qfu_name": f"{branch['branch_code']} {quote_number} / {line_number}",
                "qfu_sourceid": f"{branch['branch_code']}|SP830CA|{quote_number}|{line_number}",
                "qfu_uniquekey": f"{quote_number}_{line_number}",
                "qfu_quotenumber": quote_number,
                "qfu_linenumber": line_number,
                "qfu_soldtopartyname": first_non_empty(
                    row_value(row, headers, "Sold To Party Name", "Sold To Party Name "),
                ),
                "qfu_soldtopartycode": as_text(row_value(row, headers, "Sold-To Party")),
                "qfu_amount": total_value,
                "qfu_cssr": as_text(row_value(row, headers, "CSSR")),
                "qfu_cssrname": as_text(row_value(row, headers, "CSSR Name")),
                "qfu_description": first_non_empty(
                    row_value(row, headers, "Manufacturer Part Description"),
                    row_value(row, headers, "Additional Information"),
                ),
                "qfu_dollarconnected": converted_quote,
                "qfu_followupowner": first_non_empty(
                    row_value(row, headers, "Follow up Owner", "Follow up Owner "),
                    row_value(row, headers, "AM Name"),
                    row_value(row, headers, "CSSR Name"),
                ),
                "qfu_gppercent": round(as_float(row_value(row, headers, "GP% Rep")), 4),
                "qfu_itemnumber": as_text(row_value(row, headers, "Item Number", "Item Number ")),
                "qfu_tsr": as_text(row_value(row, headers, "AM Number")),
                "qfu_tsrname": as_text(row_value(row, headers, "AM Name")),
                "qfu_linetotal": total_value,
                "qfu_location": location,
                "qfu_manufacturerpart": as_text(row_value(row, headers, "Manufacturer Part Description")),
                "qfu_productname": as_text(row_value(row, headers, "Manufacturer Part Description")),
                "qfu_rejectionreason": rejection_reason,
                "qfu_status": status_code,
                "qfu_lastimportdate": captured_on,
                "qfu_sourcedate": iso_date(created_on),
                "qfu_branchcode": branch["branch_code"],
                "qfu_branchslug": branch["branch_slug"],
                "qfu_regionslug": branch["region_slug"],
                "qfu_sourcefamily": "SP830CA",
                "qfu_sourcefile": path.name,
                "qfu_sourceworksheet": "Daily",
            }
        )

    return sorted(line_records, key=lambda item: (item["qfu_quotenumber"] or "", item["qfu_linenumber"] or ""))


def parse_backorder_file(path, branch):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["ZBO"]
    headers, rows = read_rows(sheet, 0)
    results = []

    for row in rows:
        sales_doc = as_text(row[headers.get("Sales Doc #", -1)]) if "Sales Doc #" in headers else None
        if sales_doc is None:
            sales_doc = as_text(row[headers.get("Sales Doc _x0023_", -1)]) if "Sales Doc _x0023_" in headers else None
        if sales_doc is None:
            continue

        line = as_text(row[headers.get("Line", -1)]) if "Line" in headers else None
        on_time_date = as_date(row[headers.get("On-Time Date", -1)]) if "On-Time Date" in headers else None
        today = date.today()
        if on_time_date is None:
            days_overdue = 0
        else:
            days_overdue = max((today - on_time_date).days, 0)

        material = as_text(row[headers.get("Material", -1)]) if "Material" in headers else None
        description = as_text(row[headers.get("Material Description", -1)]) if "Material Description" in headers else None
        raw_qty_on_delivery_not_pgid = round(as_float(row[headers.get("Qty on Del Not PGI'd", -1)]) if "Qty on Del Not PGI'd" in headers else 0.0, 2)
        raw_qty_not_on_delivery = round(as_float(row[headers.get("Qty Not On Del", -1)]) if "Qty Not On Del" in headers else 0.0, 2)
        qty_on_delivery_not_pgid = max(raw_qty_on_delivery_not_pgid, 0.0)
        qty_not_on_delivery = max(raw_qty_not_on_delivery, 0.0)
        if qty_not_on_delivery <= 0 and qty_on_delivery_not_pgid <= 0:
            continue
        source_id = f"{branch['branch_code']}|ZBO|{sales_doc}|{line or '0'}"

        results.append(
            {
                "qfu_name": f"{branch['branch_code']} {sales_doc}-{line or '0'}",
                "qfu_sourceid": source_id,
                "qfu_customername": as_text(row[headers.get("Sold To Name", -1)]) if "Sold To Name" in headers else None,
                "qfu_totalvalue": round(as_float(row[headers.get("Unshipped Net Value", -1)]) if "Unshipped Net Value" in headers else 0.0, 2),
                "qfu_ontimedate": iso_date(on_time_date),
                "qfu_cssrname": as_text(row[headers.get("CSSR Name", -1)]) if "CSSR Name" in headers else None,
                "qfu_daysoverdue": days_overdue,
                "qfu_salesdocnumber": sales_doc,
                "qfu_material": material,
                "qfu_description": description,
                "qfu_quantity": round(as_float(row[headers.get("Unshipped Quantity", -1)]) if "Unshipped Quantity" in headers else 0.0, 2),
                "qfu_qtybilled": round(as_float(row[headers.get("Qty Billed", -1)]) if "Qty Billed" in headers else 0.0, 2),
                "qfu_qtyondelnotpgid": qty_on_delivery_not_pgid,
                "qfu_qtynotondel": qty_not_on_delivery,
                "qfu_branchcode": branch["branch_code"],
                "qfu_branchslug": branch["branch_slug"],
                "qfu_regionslug": branch["region_slug"],
                "qfu_sourcefamily": "ZBO",
                "qfu_sourcefile": path.name,
                "qfu_sourceline": line,
            }
        )

    return sorted(results, key=lambda item: (item["qfu_salesdocnumber"] or "", item["qfu_sourceline"] or ""))


def parse_budget_file(path, branch):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Location Summary"]
    rows = list(sheet.iter_rows(values_only=True))
    branch_rows = []
    for row in rows[2:]:
        location = as_text(row[4]) if len(row) > 4 else None
        if location == branch["branch_code"]:
            branch_rows.append(row)

    cad_row = next((row for row in branch_rows if as_text(row[6]) == "CAD"), None)
    usd_row = next((row for row in branch_rows if as_text(row[6]) == "USD"), None)
    target = as_float(cad_row[7]) if cad_row else 0.0
    cad_sales = as_float(cad_row[9]) if cad_row else 0.0
    usd_sales = as_float(usd_row[9]) if usd_row else 0.0
    actual_sales = cad_sales + usd_sales
    captured_on = datetime.fromtimestamp(path.stat().st_mtime)
    month = captured_on.month
    year = captured_on.year
    month_name = captured_on.strftime("%B")
    percent = round((actual_sales / target) * 100, 2) if target else 0.0

    return [
        {
            "qfu_name": f"{branch['branch_code']} {month_name} {year} Budget",
            "qfu_sourceid": f"{branch['branch_code']}|SA1300|{year}-{month:02d}",
            "qfu_budgetname": f"{month_name} {year} Budget",
            "qfu_actualsales": round(actual_sales, 2),
            "qfu_budgetgoal": round(target, 2),
            "qfu_percentachieved": percent,
            "qfu_lastupdated": captured_on.isoformat(),
            "qfu_cadsales": round(cad_sales, 2),
            "qfu_usdsales": round(usd_sales, 2),
            "qfu_month": month,
            "qfu_monthname": month_name,
            "qfu_year": year,
            "qfu_sourcefile": path.name,
            "qfu_branchcode": branch["branch_code"],
            "qfu_branchslug": branch["branch_slug"],
            "qfu_regionslug": branch["region_slug"],
            "qfu_sourcefamily": "SA1300",
        }
    ]


def build_summary(branch, quotes, backorders, budgets, grace_days):
    today = date.today()
    grace_cutoff = today - timedelta(days=max(0, grace_days - 1))
    last_30_cutoff = today - timedelta(days=30)
    month_start = date(today.year, today.month, 1)
    next_month = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)

    open_quotes = [record for record in quotes if int(record["qfu_status"]) == 1]
    overdue_quotes = []
    for record in open_quotes:
        source_date = as_date(record["qfu_sourcedate"])
        if source_date is not None and source_date < grace_cutoff:
            overdue_quotes.append(record)

    quotes_last_30 = []
    for record in quotes:
        source_date = as_date(record["qfu_sourcedate"])
        if source_date is not None and source_date >= last_30_cutoff:
            quotes_last_30.append(record)

    won_quotes = [record for record in quotes_last_30 if int(record["qfu_status"]) == 2]
    lost_quotes = [record for record in quotes_last_30 if int(record["qfu_status"]) == 3]

    overdue_backorders = [record for record in backorders if int(record["qfu_daysoverdue"]) > 0]
    current_month_backorders = []
    current_month_late = []
    for record in backorders:
        on_time = as_date(record["qfu_ontimedate"])
        if on_time is None:
            continue
        if month_start <= on_time < next_month:
            current_month_backorders.append(record)
            if int(record["qfu_daysoverdue"]) > 0:
                current_month_late.append(record)

    latest_budget = budgets[0] if budgets else None
    sync_label = datetime.now().isoformat()

    return {
        "qfu_name": f"{branch['branch_code']} Daily Summary {today.isoformat()}",
        "qfu_sourceid": f"{branch['branch_code']}|summary|{today.isoformat()}",
        "qfu_branchcode": branch["branch_code"],
        "qfu_branchslug": branch["branch_slug"],
        "qfu_regionslug": branch["region_slug"],
        "qfu_summarydate": today.isoformat(),
        "qfu_openquotes": len(open_quotes),
        "qfu_overduequotes": len(overdue_quotes),
        "qfu_duetoday": 0,
        "qfu_unscheduledold": len(overdue_quotes),
        "qfu_openquotevalue": round(sum(record["qfu_amount"] for record in open_quotes), 2),
        "qfu_quoteslast30days": len(quotes_last_30),
        "qfu_quoteswon30days": len(won_quotes),
        "qfu_quoteslost30days": len(lost_quotes),
        "qfu_quotesopen30days": max(0, len(quotes_last_30) - len(won_quotes) - len(lost_quotes)),
        "qfu_avgquotevalue30days": round(sum(record["qfu_amount"] for record in quotes_last_30) / len(quotes_last_30), 2) if quotes_last_30 else 0.0,
        "qfu_backordercount": len(backorders),
        "qfu_overduebackordercount": len(overdue_backorders),
        "qfu_currentmonthforecastvalue": round(sum(record["qfu_totalvalue"] for record in current_month_backorders), 2),
        "qfu_currentmonthlatevalue": round(sum(record["qfu_totalvalue"] for record in current_month_late), 2),
        "qfu_allbackordersvalue": round(sum(record["qfu_totalvalue"] for record in backorders), 2),
        "qfu_overduebackordersvalue": round(sum(record["qfu_totalvalue"] for record in overdue_backorders), 2),
        "qfu_budgetactual": round(latest_budget["qfu_actualsales"], 2) if latest_budget else 0.0,
        "qfu_budgettarget": round(latest_budget["qfu_budgetgoal"], 2) if latest_budget else 0.0,
        "qfu_budgetpace": round(latest_budget["qfu_percentachieved"], 2) if latest_budget else 0.0,
        "qfu_cadsales": round(latest_budget["qfu_cadsales"], 2) if latest_budget else 0.0,
        "qfu_usdsales": round(latest_budget["qfu_usdsales"], 2) if latest_budget else 0.0,
        "qfu_lastcalculatedon": sync_label,
    }


def build_branch_batches(branch, quotes, quote_lines, backorders, budgets):
    return [
        {
            "qfu_name": f"{branch['branch_code']} Quote Workbook Import",
            "qfu_sourceid": f"{branch['branch_code']}|batch|SP830CA",
            "qfu_branchcode": branch["branch_code"],
            "qfu_branchslug": branch["branch_slug"],
            "qfu_regionslug": branch["region_slug"],
            "qfu_sourcefamily": "SP830CA",
            "qfu_sourcefilename": quotes["file_name"],
            "qfu_status": "ready",
            "qfu_insertedcount": len(quotes["records"]) + len(quote_lines["records"]),
            "qfu_updatedcount": 0,
            "qfu_startedon": quotes["captured_on"],
            "qfu_completedon": quotes["captured_on"],
            "qfu_triggerflow": "Controlled workbook seed",
            "qfu_notes": f"Seeded {len(quotes['records'])} quote headers and {len(quote_lines['records'])} quote lines from {quotes['file_name']} for {branch['branch_code']}.",
        },
        {
            "qfu_name": f"{branch['branch_code']} Backorder Workbook Import",
            "qfu_sourceid": f"{branch['branch_code']}|batch|ZBO",
            "qfu_branchcode": branch["branch_code"],
            "qfu_branchslug": branch["branch_slug"],
            "qfu_regionslug": branch["region_slug"],
            "qfu_sourcefamily": "ZBO",
            "qfu_sourcefilename": backorders["file_name"],
            "qfu_status": "ready",
            "qfu_insertedcount": len(backorders["records"]),
            "qfu_updatedcount": 0,
            "qfu_startedon": backorders["captured_on"],
            "qfu_completedon": backorders["captured_on"],
            "qfu_triggerflow": "Controlled workbook seed",
            "qfu_notes": f"Seeded from {backorders['file_name']} for {branch['branch_code']}.",
        },
        {
            "qfu_name": f"{branch['branch_code']} Budget Workbook Import",
            "qfu_sourceid": f"{branch['branch_code']}|batch|SA1300",
            "qfu_branchcode": branch["branch_code"],
            "qfu_branchslug": branch["branch_slug"],
            "qfu_regionslug": branch["region_slug"],
            "qfu_sourcefamily": "SA1300",
            "qfu_sourcefilename": budgets["file_name"],
            "qfu_status": "ready",
            "qfu_insertedcount": len(budgets["records"]) + len(budgets.get("ops_daily_records", [])),
            "qfu_updatedcount": 0,
            "qfu_startedon": budgets["captured_on"],
            "qfu_completedon": budgets["captured_on"],
            "qfu_triggerflow": "Controlled workbook seed",
            "qfu_notes": f"Seeded from {budgets['file_name']} for {branch['branch_code']} with {len(budgets.get('ops_daily_records', []))} SA1300 ops daily rows.",
        },
    ]


def parse_branch(example_root, branch_code, grace_days):
    branch = BRANCHES[branch_code]
    branch_root = Path(example_root, branch_code)
    quote_path = next(branch_root.glob("*Quote Follow Up Report*.xlsx"))
    backorder_path = next(branch_root.glob("*ZBO*.xlsx"))
    budget_path = next(branch_root.glob("SA1300*.xlsx"))

    quote_records = parse_quote_file(quote_path, branch)
    quote_line_records = parse_quote_lines(quote_path, branch)
    backorder_records = parse_backorder_file(backorder_path, branch)
    budget_records = parse_budget_file(budget_path, branch)
    ops_daily_payload = parse_ops_daily(budget_path, branch)
    for record in budget_records:
        record["qfu_opsdailycadjson"] = compact_json(ops_daily_payload["cad_rows"])
        record["qfu_opsdailyusdjson"] = compact_json(ops_daily_payload["usd_rows"])
        record["qfu_opsdailyasof"] = ops_daily_payload["latest_billing_day"]
    summary_record = build_summary(branch, quote_records, backorder_records, budget_records, grace_days)

    quotes_payload = {
        "file_name": quote_path.name,
        "captured_on": datetime.fromtimestamp(quote_path.stat().st_mtime).isoformat(),
        "records": quote_records,
    }
    quote_lines_payload = {
        "file_name": quote_path.name,
        "captured_on": datetime.fromtimestamp(quote_path.stat().st_mtime).isoformat(),
        "records": quote_line_records,
    }
    backorders_payload = {
        "file_name": backorder_path.name,
        "captured_on": datetime.fromtimestamp(backorder_path.stat().st_mtime).isoformat(),
        "records": backorder_records,
    }
    budgets_payload = {
        "file_name": budget_path.name,
        "captured_on": datetime.fromtimestamp(budget_path.stat().st_mtime).isoformat(),
        "records": budget_records,
        "ops_daily_records": ops_daily_payload["records"],
    }

    return {
        "branch": branch,
        "quotes": quotes_payload,
        "quote_lines": quote_lines_payload,
        "backorders": backorders_payload,
        "budgets": budgets_payload,
        "ops_daily": ops_daily_payload,
        "summary": summary_record,
        "batches": build_branch_batches(branch, quotes_payload, quote_lines_payload, backorders_payload, budgets_payload),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--example-root", default="example")
    parser.add_argument("--output", required=True)
    parser.add_argument("--grace-days", type=int, default=3)
    args = parser.parse_args()

    branches = [parse_branch(args.example_root, code, args.grace_days) for code in sorted(BRANCHES.keys())]
    payload = {
        "generated_on": datetime.now().isoformat(),
        "region": {
            "region_slug": "southern-alberta",
            "region_name": "Southern Alberta",
        },
        "branches": branches,
    }

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()

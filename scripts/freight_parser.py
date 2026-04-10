import argparse
import datetime as dt
import json
import math
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl
import xlrd


FREIGHT_HIGH_VALUE_THRESHOLD = 250.0

APPLIED_MARKERS = (
    "APPLIED",
    "APPLIED INDUSTRIAL",
    "APPLIED INDUSTRIAL TECHNOLOG",
    "APPLIED INDUSTRIAL TECH",
    "APPLIED INDL",
    "APPLIED LETHBRIDGE",
    "APPLIED CALGARY",
    "APPLIED MEDICINE HAT",
)

FAMILY_LABELS = {
    "FREIGHT_REDWOOD": "Applied / Redwood",
    "FREIGHT_LOOMIS_F15": "Loomis",
    "FREIGHT_PUROLATOR_F07": "Purolator",
    "FREIGHT_UPS_F06": "UPS",
}


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return "" if text.lower() == "nan" else text


def normalize_slug_component(value, max_length=72):
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    if not text:
        return "none"
    return text[:max_length].strip("-") or "none"


def to_float(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def round_money(value):
    if value is None:
        return None
    return round(float(value), 2)


def parse_excel_date(value, datemode=None):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if datemode is None:
            return None
        try:
            parsed = xlrd.xldate.xldate_as_datetime(float(value), datemode)
        except Exception:
            return None
        return parsed.date().isoformat()

    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("T", " ")
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
    ):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return dt.datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return None


def choose_preferred(values, placeholders=None):
    placeholders = set(placeholders or ("", "???", "nan", "none"))
    candidates = []
    for value in values:
        text = normalize_text(value)
        if not text or text.lower() in placeholders:
            continue
        candidates.append(text)
    if not candidates:
        return ""
    return sorted(candidates, key=lambda item: (len(item), item), reverse=True)[0]


def has_applied_marker(text):
    upper = normalize_text(text).upper()
    if not upper:
        return False
    return any(marker in upper for marker in APPLIED_MARKERS)


def infer_direction(sender, destination):
    sender_internal = has_applied_marker(sender)
    destination_internal = has_applied_marker(destination)
    if sender_internal and destination_internal:
        return "Internal"
    if sender_internal:
        return "Outbound"
    return "Inbound"


def determine_priority_band(total_amount):
    amount = to_float(total_amount) or 0.0
    return "High Value" if amount >= FREIGHT_HIGH_VALUE_THRESHOLD else "Standard"


def parse_charge_breakdown(text):
    raw_text = normalize_text(text)
    totals = {
        "freight": 0.0,
        "fuel": 0.0,
        "gst": 0.0,
        "hst": 0.0,
        "qst": 0.0,
        "accessorial": 0.0,
    }
    if not raw_text:
        return {
            "freightamount": None,
            "fuelamount": None,
            "gstamount": None,
            "hstamount": None,
            "qstamount": None,
            "taxamount": None,
            "accessorialamount": None,
            "accessorialpresent": False,
            "totalamount": None,
            "chargebreakdowntext": "",
        }

    parts = [segment.strip() for segment in raw_text.split("|") if segment.strip()]
    for part in parts:
        match = re.match(r"^\s*([+-]?\d+(?:\.\d+)?)\s+(.*)$", part)
        if not match:
            continue
        amount = float(match.group(1))
        label = match.group(2).strip().lower()
        if "gst" in label:
            totals["gst"] += amount
        elif "hst" in label:
            totals["hst"] += amount
        elif "qst" in label:
            totals["qst"] += amount
        elif "fuel" in label:
            totals["fuel"] += amount
        elif any(keyword in label for keyword in ("handling", "declared value", "delivery confirmation", "invoice service", "special", "accessorial", "surcharge")):
            totals["accessorial"] += amount
        else:
            totals["freight"] += amount

    tax_amount = totals["gst"] + totals["hst"] + totals["qst"]
    total_amount = totals["freight"] + totals["fuel"] + totals["accessorial"] + tax_amount
    return {
        "freightamount": round_money(totals["freight"]) if total_amount else None,
        "fuelamount": round_money(totals["fuel"]) if total_amount else None,
        "gstamount": round_money(totals["gst"]) if totals["gst"] else None,
        "hstamount": round_money(totals["hst"]) if totals["hst"] else None,
        "qstamount": round_money(totals["qst"]) if totals["qst"] else None,
        "taxamount": round_money(tax_amount) if tax_amount else None,
        "accessorialamount": round_money(totals["accessorial"]) if totals["accessorial"] else None,
        "accessorialpresent": totals["accessorial"] > 0,
        "totalamount": round_money(total_amount) if total_amount else None,
        "chargebreakdowntext": " | ".join(parts),
    }


def build_source_id(branch_code, source_family, control_number, invoice_number, tracking_number, reference_value, service_code=""):
    parts = [
        normalize_slug_component(branch_code, 16),
        normalize_slug_component(source_family, 40),
        normalize_slug_component(control_number, 48),
        normalize_slug_component(invoice_number, 48),
        normalize_slug_component(tracking_number, 72),
        normalize_slug_component(reference_value, 96),
        normalize_slug_component(service_code, 32),
    ]
    return "|".join(parts)


def dump_raw_payload(payload):
    return json.dumps(payload, ensure_ascii=True, default=str, separators=(",", ":"))


def read_xlsx_rows(path):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [normalize_text(value) for value in rows[0]]
    result = []
    for raw_row in rows[1:]:
        mapped = {header[index]: raw_row[index] for index in range(len(header))}
        if not any(normalize_text(value) for value in mapped.values()):
            continue
        result.append(mapped)
    return result


def read_xls_sheet_rows(path, preferred_sheet="Sheet2"):
    workbook = xlrd.open_workbook(path)
    if preferred_sheet in workbook.sheet_names():
        sheet = workbook.sheet_by_name(preferred_sheet)
    else:
        sheet = workbook.sheet_by_index(0)

    header_row = None
    for row_index in range(min(sheet.nrows, 10)):
        row_values = [normalize_text(value) for value in sheet.row_values(row_index)]
        if sum(1 for value in row_values if value) >= 8:
            header_row = row_index
            header = row_values
            break
    if header_row is None:
        return [], workbook.datemode

    rows = []
    for row_index in range(header_row + 1, sheet.nrows):
        row_values = sheet.row_values(row_index)
        row = {header[index]: row_values[index] for index in range(len(header))}
        if not any(normalize_text(value) for value in row.values()):
            continue
        if normalize_text(row.get(header[0])).upper() == "TOTAL":
            continue
        rows.append(row)
    return rows, workbook.datemode


def build_redwood_records(rows, branch_code, branch_slug, region_slug, source_filename, import_batch_id):
    records = []
    for row in rows:
        invoice_number = normalize_text(row.get("Invoice Number"))
        if not invoice_number:
            continue
        primary_reference = normalize_text(row.get("Primary Reference"))
        ref1 = normalize_text(row.get("Ref1"))
        pro_number = normalize_text(row.get("PRO"))
        control_number = normalize_text(row.get("BOL")) or normalize_text(row.get("Check Number")) or normalize_text(row.get("Extract Number"))
        source_id = build_source_id(
            branch_code,
            "FREIGHT_REDWOOD",
            control_number or branch_code,
            invoice_number,
            pro_number or normalize_text(row.get("Primary Reference")),
            ref1 or primary_reference,
            normalize_text(row.get("Invoice SCAC")),
        )
        gst_amount = to_float(row.get("GST"))
        hst_amount = to_float(row.get("HST"))
        qst_amount = to_float(row.get("QST"))
        accessorial_amount = sum(
            filter(
                None,
                [
                    to_float(row.get("Accessorial - CTS")),
                    to_float(row.get("Accessorial - MSC")),
                    to_float(row.get("ACC 1 Amount")),
                    to_float(row.get("ACC 2 Amount")),
                    to_float(row.get("ACC 3 Amount")),
                    to_float(row.get("ACC 4 Amount")),
                ],
            )
        ) or None
        total_amount = to_float(row.get("Total Invoice Amount Tax Included"))
        freight_amount = to_float(row.get("Freight Cost Only")) or to_float(row.get("Line Haul"))
        fuel_amount = to_float(row.get("Fuel"))
        charge_parts = []
        for label, value in (
            ("Freight", freight_amount),
            ("Fuel", fuel_amount),
            ("GST", gst_amount),
            ("HST", hst_amount),
            ("QST", qst_amount),
            ("Accessorial", accessorial_amount),
        ):
            if value is not None:
                charge_parts.append(f"{value:.2f} {label}")
        if normalize_text(row.get("Least Cost Carrier")):
            charge_parts.append(
                "Least Cost Carrier: "
                + normalize_text(row.get("Least Cost Carrier"))
                + " | Normalized Total: "
                + (f"{to_float(row.get('Least Cost Normalized Total')):.2f}" if to_float(row.get("Least Cost Normalized Total")) is not None else "n/a")
            )

        sender = ", ".join(
            filter(
                None,
                [
                    normalize_text(row.get("Origin Name")),
                    normalize_text(row.get("Origin Addr1")),
                    normalize_text(row.get("Origin City")),
                    normalize_text(row.get("Origin State")),
                    normalize_text(row.get("Origin Zip")),
                ],
            )
        )
        destination = ", ".join(
            filter(
                None,
                [
                    normalize_text(row.get("Dest Name")),
                    normalize_text(row.get("Dest Addr1")),
                    normalize_text(row.get("Dest City")),
                    normalize_text(row.get("Dest State")),
                    normalize_text(row.get("Dest Zip")),
                ],
            )
        )
        tax_amount = sum(filter(None, [gst_amount, hst_amount, qst_amount])) or None
        total_for_priority = total_amount if total_amount is not None else (
            (freight_amount or 0.0) + (fuel_amount or 0.0) + (accessorial_amount or 0.0) + (tax_amount or 0.0)
        )
        record = OrderedDict(
            qfu_name=f"{branch_code} Freight {invoice_number}",
            qfu_sourceid=source_id,
            qfu_branchcode=branch_code,
            qfu_branchslug=branch_slug,
            qfu_regionslug=region_slug,
            qfu_sourcefamily="FREIGHT_REDWOOD",
            qfu_sourcefilename=source_filename,
            qfu_sourcecarrier=normalize_text(row.get("Invoice Carrier")) or "Applied / Redwood",
            qfu_importbatchid=import_batch_id,
            qfu_trackingnumber=pro_number or normalize_text(row.get("BOL")) or primary_reference,
            qfu_pronumber=pro_number,
            qfu_invoicenumber=invoice_number,
            qfu_controlnumber=control_number,
            qfu_reference=ref1 or primary_reference,
            qfu_shipdate=parse_excel_date(row.get("Actual Ship")) or parse_excel_date(row.get("Target Ship (Early)")),
            qfu_invoicedate=parse_excel_date(row.get("Extract Date")) or parse_excel_date(row.get("Check Date")),
            qfu_closedate=parse_excel_date(row.get("Actual Delivery")) or parse_excel_date(row.get("Target Delivery (Early)")),
            qfu_billtype="",
            qfu_service=normalize_text(row.get("Carrier Mode")),
            qfu_servicecode=normalize_text(row.get("Invoice SCAC")),
            qfu_sender=sender,
            qfu_destination=destination,
            qfu_zone="",
            qfu_actualweight=round_money(to_float(row.get("Weight"))),
            qfu_billedweight=round_money(to_float(row.get("Invoice Weight"))),
            qfu_quantity=round_money(to_float(row.get("Quantity"))),
            qfu_totalamount=round_money(total_amount if total_amount is not None else total_for_priority),
            qfu_freightamount=round_money(freight_amount),
            qfu_fuelamount=round_money(fuel_amount),
            qfu_taxamount=round_money(tax_amount),
            qfu_gstamount=round_money(gst_amount),
            qfu_hstamount=round_money(hst_amount),
            qfu_qstamount=round_money(qst_amount),
            qfu_accessorialamount=round_money(accessorial_amount),
            qfu_accessorialpresent=bool((accessorial_amount or 0) > 0 or normalize_text(row.get("ACC 1")) or normalize_text(row.get("ACC 2")) or normalize_text(row.get("ACC 3")) or normalize_text(row.get("ACC 4")) or normalize_text(row.get("ACC 5+"))),
            qfu_unrealizedsavings=round_money(to_float(row.get("Normalized Unrealized Savings"))),
            qfu_chargebreakdowntext=" | ".join(charge_parts),
            qfu_direction=infer_direction(sender, destination),
            qfu_status="Open",
            qfu_priorityband=determine_priority_band(total_for_priority),
            qfu_ownername="",
            qfu_owneridentifier="",
            qfu_claimedon=None,
            qfu_comment="",
            qfu_commentupdatedon=None,
            qfu_commentupdatedbyname="",
            qfu_lastactivityon=None,
            qfu_isarchived=False,
            qfu_archivedon=None,
            qfu_lastseenon=None,
            qfu_rawrowjson=dump_raw_payload(row),
        )
        records.append(record)
    return records


def build_carrier_record(group_key, rows, branch_code, branch_slug, region_slug, source_filename, import_batch_id, source_family, datemode):
    first = rows[0]
    invoice_number = choose_preferred([row.get("Invoice") for row in rows])
    tracking_number = choose_preferred([row.get("Tracking") for row in rows])
    reference_value = choose_preferred([row.get("Reference") for row in rows])
    control_number = choose_preferred([row.get("Shipper") for row in rows])
    sender = choose_preferred([row.get("Sender") for row in rows])
    destination = choose_preferred([row.get("Destination") for row in rows])
    service = choose_preferred([row.get("Service") for row in rows])
    service_code = choose_preferred([row.get("Code") for row in rows])
    ship_date = choose_preferred([parse_excel_date(row.get("Ship Date"), datemode) for row in rows])
    invoice_date = choose_preferred([parse_excel_date(row.get("Invoice Date"), datemode) for row in rows])
    close_date = choose_preferred([parse_excel_date(row.get("Close Date"), datemode) for row in rows])
    bill_type = choose_preferred([row.get("Bill Type") for row in rows])
    zone = choose_preferred([row.get("Zone") for row in rows])
    source_carrier = FAMILY_LABELS[source_family]
    actual_weight = max([to_float(row.get("Act Wgt")) or 0.0 for row in rows] or [0.0]) or None
    billed_weight = max([to_float(row.get("Bill Wgt")) or 0.0 for row in rows] or [0.0]) or None
    quantity = max([to_float(row.get("Qty")) or 0.0 for row in rows] or [0.0]) or None

    billed_amounts = [to_float(row.get("Billed Charges")) for row in rows]
    billed_total = sum(value for value in billed_amounts if value is not None)
    parsed_parts = [parse_charge_breakdown(row.get("Charge Description With Extended Bill Amount") or row.get("Charge Description")) for row in rows]
    parsed_total = sum(part.get("totalamount") or 0.0 for part in parsed_parts)
    gst_from_columns = sum(filter(None, [to_float(item.get("GST Amount")) for item in rows] + [to_float(item.get("GST Extended Amount")) for item in rows])) or None
    hst_from_columns = sum(filter(None, [to_float(item.get("HST Amount")) for item in rows] + [to_float(item.get("HST Extended Amount")) for item in rows])) or None
    qst_from_columns = sum(filter(None, [to_float(item.get("QST Amount")) for item in rows] + [to_float(item.get("QST Extended Amount")) for item in rows])) or None
    gst_amount = gst_from_columns if gst_from_columns is not None else (sum(filter(None, [part.get("gstamount") for part in parsed_parts])) or None)
    hst_amount = hst_from_columns if hst_from_columns is not None else (sum(filter(None, [part.get("hstamount") for part in parsed_parts])) or None)
    qst_amount = qst_from_columns if qst_from_columns is not None else (sum(filter(None, [part.get("qstamount") for part in parsed_parts])) or None)
    fuel_amount = sum(filter(None, [part.get("fuelamount") for part in parsed_parts])) or None
    freight_amount = sum(filter(None, [part.get("freightamount") for part in parsed_parts])) or None
    accessorial_amount = sum(filter(None, [part.get("accessorialamount") for part in parsed_parts])) or None
    tax_amount = sum(filter(None, [gst_amount, hst_amount, qst_amount])) or None
    total_amount = billed_total if billed_total > 0 else parsed_total
    if total_amount <= 0 and any(value is not None for value in (freight_amount, fuel_amount, accessorial_amount, tax_amount)):
        total_amount = (freight_amount or 0.0) + (fuel_amount or 0.0) + (accessorial_amount or 0.0) + (tax_amount or 0.0)

    descriptions = []
    for row in rows:
        for field_name in ("Charge Description With Extended Bill Amount", "Charge Description", "Desc"):
            value = normalize_text(row.get(field_name))
            if value and value not in descriptions:
                descriptions.append(value)
    chargebreakdowntext = " || ".join(descriptions)
    accessorial_present = bool((accessorial_amount or 0.0) > 0.0 or choose_preferred([row.get("Accessorials present?") for row in rows]).lower() in {"y", "yes", "true"})
    source_id = build_source_id(
        branch_code,
        source_family,
        control_number or branch_code,
        invoice_number,
        tracking_number,
        reference_value,
        service_code or service,
    )

    merged_payload = {
        "group_key": list(group_key),
        "source_rows": rows,
    }

    return OrderedDict(
        qfu_name=f"{branch_code} Freight {invoice_number or tracking_number}",
        qfu_sourceid=source_id,
        qfu_branchcode=branch_code,
        qfu_branchslug=branch_slug,
        qfu_regionslug=region_slug,
        qfu_sourcefamily=source_family,
        qfu_sourcefilename=source_filename,
        qfu_sourcecarrier=source_carrier,
        qfu_importbatchid=import_batch_id,
        qfu_trackingnumber=tracking_number,
        qfu_pronumber="",
        qfu_invoicenumber=invoice_number,
        qfu_controlnumber=control_number,
        qfu_reference=reference_value,
        qfu_shipdate=ship_date,
        qfu_invoicedate=invoice_date,
        qfu_closedate=close_date,
        qfu_billtype=bill_type,
        qfu_service=service,
        qfu_servicecode=service_code,
        qfu_sender=sender,
        qfu_destination=destination,
        qfu_zone=zone,
        qfu_actualweight=round_money(actual_weight),
        qfu_billedweight=round_money(billed_weight),
        qfu_quantity=round_money(quantity),
        qfu_totalamount=round_money(total_amount) if total_amount else None,
        qfu_freightamount=round_money(freight_amount),
        qfu_fuelamount=round_money(fuel_amount),
        qfu_taxamount=round_money(tax_amount),
        qfu_gstamount=round_money(gst_amount),
        qfu_hstamount=round_money(hst_amount),
        qfu_qstamount=round_money(qst_amount),
        qfu_accessorialamount=round_money(accessorial_amount),
        qfu_accessorialpresent=accessorial_present,
        qfu_unrealizedsavings=None,
        qfu_chargebreakdowntext=chargebreakdowntext,
        qfu_direction=infer_direction(sender, destination),
        qfu_status="Open",
        qfu_priorityband=determine_priority_band(total_amount),
        qfu_ownername="",
        qfu_owneridentifier="",
        qfu_claimedon=None,
        qfu_comment="",
        qfu_commentupdatedon=None,
        qfu_commentupdatedbyname="",
        qfu_lastactivityon=None,
        qfu_isarchived=False,
        qfu_archivedon=None,
        qfu_lastseenon=None,
        qfu_rawrowjson=dump_raw_payload(merged_payload),
    )


def build_grouped_carrier_records(rows, branch_code, branch_slug, region_slug, source_filename, import_batch_id, source_family, datemode):
    groups = OrderedDict()
    for row in rows:
        key = (
            normalize_text(row.get("Invoice")),
            normalize_text(row.get("Tracking")),
            normalize_text(row.get("Reference")),
            normalize_text(row.get("Shipper")),
        )
        groups.setdefault(key, []).append(row)

    records = [
        build_carrier_record(
            group_key=group_key,
            rows=group_rows,
            branch_code=branch_code,
            branch_slug=branch_slug,
            region_slug=region_slug,
            source_filename=source_filename,
            import_batch_id=import_batch_id,
            source_family=source_family,
            datemode=datemode,
        )
        for group_key, group_rows in groups.items()
    ]
    return records, len(rows) - len(records)


def infer_source_family(path):
    file_name = path.name.lower()
    if "loomis" in file_name:
        return "FREIGHT_LOOMIS_F15"
    if "purolator" in file_name:
        return "FREIGHT_PUROLATOR_F07"
    if "ups" in file_name:
        return "FREIGHT_UPS_F06"
    if path.suffix.lower() == ".xlsx":
        return "FREIGHT_REDWOOD"
    raise ValueError(f"Unable to infer freight source family for {path}")


def parse_freight_file(input_path, branch_code, branch_slug, region_slug, source_family=None, source_filename=None, import_batch_id=""):
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(path)
    source_family = normalize_text(source_family) or infer_source_family(path)
    source_filename = normalize_text(source_filename) or path.name

    if source_family == "FREIGHT_REDWOOD":
        rows = read_xlsx_rows(path)
        records = build_redwood_records(
            rows=rows,
            branch_code=branch_code,
            branch_slug=branch_slug,
            region_slug=region_slug,
            source_filename=source_filename,
            import_batch_id=import_batch_id,
        )
        collapsed = 0
    else:
        rows, datemode = read_xls_sheet_rows(path)
        records, collapsed = build_grouped_carrier_records(
            rows=rows,
            branch_code=branch_code,
            branch_slug=branch_slug,
            region_slug=region_slug,
            source_filename=source_filename,
            import_batch_id=import_batch_id,
            source_family=source_family,
            datemode=datemode,
        )

    return OrderedDict(
        input_path=str(path),
        source_family=source_family,
        source_label=FAMILY_LABELS.get(source_family, source_family),
        source_filename=source_filename,
        branch_code=branch_code,
        branch_slug=branch_slug,
        region_slug=region_slug,
        input_row_count=len(rows),
        normalized_record_count=len(records),
        collapsed_group_rows=collapsed,
        high_value_threshold=FREIGHT_HIGH_VALUE_THRESHOLD,
        records=records,
    )


def cli():
    parser = argparse.ArgumentParser(description="Normalize weekly freight reports into QFU freight worklist rows.")
    parser.add_argument("--input", required=True, help="Path to the freight workbook.")
    parser.add_argument("--branch-code", required=True)
    parser.add_argument("--branch-slug", required=True)
    parser.add_argument("--region-slug", required=True)
    parser.add_argument("--source-family", default="")
    parser.add_argument("--source-filename", default="")
    parser.add_argument("--import-batch-id", default="")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    payload = parse_freight_file(
        input_path=args.input,
        branch_code=args.branch_code,
        branch_slug=args.branch_slug,
        region_slug=args.region_slug,
        source_family=args.source_family,
        source_filename=args.source_filename,
        import_batch_id=args.import_batch_id,
    )
    Path(args.output).write_text(json.dumps(payload, ensure_ascii=True, indent=2, default=str), encoding="utf-8")


if __name__ == "__main__":
    cli()

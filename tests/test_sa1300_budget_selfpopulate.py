import json
import unittest
from email import policy
from email.parser import BytesParser
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKFLOW_ROOT = REPO_ROOT / "results" / "sapilotflows" / "src" / "Workflows"
EXAMPLE_ROOT_CANDIDATES = (
    REPO_ROOT / "example",
    REPO_ROOT.parent / "example",
)

WORKFLOW_IDS = {
    "4171": "6DB19FF3-C313-4DB6-9A57-F3335FE55558",
    "4172": "078CEA4C-84F6-4C4F-B73B-62AD838F7CAE",
    "4173": "3C2EBD80-35D9-4E3C-BDBE-70BE98A82AE6",
}

EXPECTED_MONTH_END_PLAN = {
    "4171": 945500.0,
    "4172": 388540.0,
    "4173": 411990.0,
}

EXPECTED_TARGET_RANGE = "'Location Summary'!H2:H500"
EXPECTED_RESOLVED_EXPR = (
    "@coalesce(first(outputs('Get_Budget_Goal_From_Archives')?['body/value'])?['qfu_budgetgoal'], "
    "outputs('Resolve_Budget_Goal_From_SA1300_Plan'), "
    "outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_budgetgoal'])"
)
CURRENT_MONTH_SOURCE_EXPR = (
    "concat(parameters('qfu_QFU_BranchCode'), '|SA1300|', "
    "formatDateTime(convertTimeZone(utcNow(), 'UTC', 'Mountain Standard Time'), 'yyyy-MM'))"
)
EXPECTED_PRESERVED_ACTUAL_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], "
    "variables('TotalSales'))"
)
EXPECTED_PRESERVED_CAD_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_cadsales'], "
    "variables('CADSales'))"
)
EXPECTED_PRESERVED_USD_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_usdsales'], "
    "variables('USDSales'))"
)
EXPECTED_PRESERVED_SOURCEFILE_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourcefile'], "
    "items('Apply_to_each_Attachment')?['name'])"
)
EXPECTED_PRESERVED_LASTUPDATED_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_lastupdated'], "
    "utcNow())"
)
EXPECTED_PRESERVED_CAD_OPS_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_opsdailycadjson'], "
    "string(json('[]'))), string(coalesce(body('Filter_CAD_Ops_Daily_Rows'), json('[]'))))"
)
EXPECTED_PRESERVED_USD_OPS_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_opsdailyusdjson'], "
    "string(json('[]'))), string(coalesce(outputs('Filter_USD_Ops_Daily_Rows')?['body'], json('[]'))))"
)
EXPECTED_DELETE_OPS_FOREACH_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), json('[]'), coalesce(outputs('List_Existing_Branch_Ops_Daily')?['body/value'], "
    "json('[]')))"
)
EXPECTED_APPLY_USD_FOREACH_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), json('[]'), coalesce(body('Filter_USD_Ops_Daily_Rows'), json('[]')))"
)
EXPECTED_APPLY_CAD_FOREACH_EXPR = (
    "@if(and(equals(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_sourceid'], "
    + CURRENT_MONTH_SOURCE_EXPR
    + "), greater(float(coalesce(outputs('Get_Active_Budget')?['body/value']?[0]?['qfu_actualsales'], 0)), "
    "float(coalesce(variables('TotalSales'), 0)))), json('[]'), coalesce(body('Filter_CAD_Ops_Daily_Rows'), json('[]')))"
)
EXPECTED_ACTIVE_BUDGET_SELECT = (
    "qfu_budgetid,qfu_sourceid,qfu_budgetgoal,qfu_actualsales,qfu_cadsales,qfu_usdsales,"
    "qfu_opsdailycadjson,qfu_opsdailyusdjson,qfu_sourcefile,qfu_lastupdated,qfu_month,"
    "qfu_monthname,qfu_year,qfu_fiscalyear"
)


def resolve_example_root() -> Path:
    for candidate in EXAMPLE_ROOT_CANDIDATES:
        if candidate.exists():
            return candidate

    searched = ", ".join(str(candidate) for candidate in EXAMPLE_ROOT_CANDIDATES)
    raise FileNotFoundError(f"Example workbook root not found. Checked: {searched}")


EXAMPLE_ROOT = resolve_example_root()


def branch_example_dirs(branch_code: str) -> tuple[Path, ...]:
    return (
        EXAMPLE_ROOT / branch_code,
        EXAMPLE_ROOT / "April 9" / branch_code,
    )


def resolve_sa1300_workbook_source(branch_code: str):
    for directory in branch_example_dirs(branch_code):
        if not directory.exists():
            continue
        workbook_matches = sorted(directory.glob("*SA1300*.xlsx"))
        if workbook_matches:
            return workbook_matches[0]

    for directory in branch_example_dirs(branch_code):
        if not directory.exists():
            continue
        for message_path in sorted(directory.glob("*SA1300*.eml")):
            with message_path.open("rb") as handle:
                message = BytesParser(policy=policy.default).parse(handle)
            for attachment in message.iter_attachments():
                filename = (attachment.get_filename() or "").lower()
                content_type = (attachment.get_content_type() or "").lower()
                if filename.endswith(".xlsx") or "spreadsheet" in content_type:
                    payload = attachment.get_payload(decode=True)
                    if payload:
                        return BytesIO(payload)

    searched = ", ".join(str(directory) for directory in branch_example_dirs(branch_code))
    raise FileNotFoundError(f"SA1300 workbook or workbook email attachment not found for {branch_code}. Checked: {searched}")


def workflow_root_actions(branch_code: str) -> dict:
    workflow_id = WORKFLOW_IDS[branch_code]
    path = WORKFLOW_ROOT / f"{branch_code}-Budget-Update-SA1300-{workflow_id}.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload["properties"]["definition"]["actions"]["Apply_to_each_Attachment"]["actions"][
        "Condition_Is_SA1300_File"
    ]["actions"]


def workflow_guard_actions(branch_code: str) -> dict:
    return workflow_root_actions(branch_code)["Guard_Budget_Row_Limit"]["actions"]


def workflow_same_month_update(branch_code: str) -> dict:
    actions = workflow_guard_actions(branch_code)
    return actions["Condition_Check_Month_Changed"]["else"]["actions"]["Condition_Budget_Exists_Same_Month"][
        "actions"
    ]["Update_Current_Month_Budget"]["inputs"]["parameters"]["item"]


def month_end_plan_from_example(branch_code: str) -> float:
    workbook_source = resolve_sa1300_workbook_source(branch_code)
    workbook = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        sheet = workbook["Location Summary"]
        for row in sheet.iter_rows(values_only=True):
            if len(row) > 7 and str(row[4]).strip() == branch_code and str(row[6]).strip() == "CAD":
                return float(row[7])
    finally:
        workbook.close()
    raise AssertionError(f"Month-End Plan CAD row not found for {branch_code}")


class Sa1300BudgetSelfPopulateTests(unittest.TestCase):
    def test_example_workbooks_expose_month_end_plan_sales(self) -> None:
        for branch_code, expected_value in EXPECTED_MONTH_END_PLAN.items():
            with self.subTest(branch=branch_code):
                actual_value = month_end_plan_from_example(branch_code)
                self.assertAlmostEqual(actual_value, expected_value, places=2)

    def test_budget_flow_extracts_month_end_plan_target_table(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                actions = workflow_root_actions(branch_code)
                self.assertIn("Create_Budget_Target_Table", actions)
                self.assertEqual(
                    actions["Create_Budget_Target_Table"]["inputs"]["parameters"]["table/Range"],
                    EXPECTED_TARGET_RANGE,
                )
                self.assertEqual(
                    actions["Resolve_Budget_Goal_From_SA1300_Plan"]["runAfter"],
                    {"Filter_Budget_Target_Rows": ["Succeeded"]},
                )
                self.assertEqual(
                    actions["Guard_Budget_Row_Limit"]["runAfter"],
                    {"Resolve_Budget_Goal_From_SA1300_Plan": ["Succeeded"]},
                )

    def test_budget_flow_resolves_goal_from_archive_then_sa1300_then_active_budget(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                actions = workflow_guard_actions(branch_code)
                self.assertEqual(actions["Get_Active_Budget"]["runAfter"], {"Get_Budget_Goal_From_Archives": ["Succeeded"]})
                ensure_expr = actions["Ensure_Budget_Goal_Found"]["expression"]
                self.assertEqual(
                    ensure_expr,
                    {"and": [{"not": {"equals": [EXPECTED_RESOLVED_EXPR, None]}}]},
                )
                self.assertEqual(
                    actions["Ensure_Budget_Goal_Found"]["runAfter"],
                    {"Get_Active_Budget": ["Succeeded"]},
                )

    def test_budget_rows_write_resolved_goal_output(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                actions = workflow_guard_actions(branch_code)
                month_changed = actions["Condition_Check_Month_Changed"]
                same_month_update = workflow_same_month_update(branch_code)
                self.assertEqual(same_month_update["qfu_budgetgoal"], EXPECTED_RESOLVED_EXPR)
                self.assertEqual(same_month_update["qfu_budgetamount"], EXPECTED_RESOLVED_EXPR)

                create_first = month_changed["else"]["actions"]["Condition_Budget_Exists_Same_Month"]["else"][
                    "actions"
                ]["Create_First_Budget_Record"]["inputs"]["parameters"]
                self.assertEqual(create_first["item/qfu_budgetamount"], EXPECTED_RESOLVED_EXPR)

    def test_budget_flow_selects_current_budget_fields_for_rollback_guard(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                actions = workflow_guard_actions(branch_code)
                self.assertEqual(
                    actions["Get_Active_Budget"]["inputs"]["parameters"]["$select"],
                    EXPECTED_ACTIVE_BUDGET_SELECT,
                )

    def test_same_month_budget_update_preserves_higher_live_actuals(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                same_month_update = workflow_same_month_update(branch_code)
                self.assertEqual(same_month_update["qfu_actualsales"], EXPECTED_PRESERVED_ACTUAL_EXPR)
                self.assertEqual(same_month_update["qfu_cadsales"], EXPECTED_PRESERVED_CAD_EXPR)
                self.assertEqual(same_month_update["qfu_usdsales"], EXPECTED_PRESERVED_USD_EXPR)
                self.assertEqual(same_month_update["qfu_sourcefile"], EXPECTED_PRESERVED_SOURCEFILE_EXPR)
                self.assertEqual(same_month_update["qfu_lastupdated"], EXPECTED_PRESERVED_LASTUPDATED_EXPR)
                self.assertEqual(same_month_update["qfu_opsdailycadjson"], EXPECTED_PRESERVED_CAD_OPS_EXPR)
                self.assertEqual(same_month_update["qfu_opsdailyusdjson"], EXPECTED_PRESERVED_USD_OPS_EXPR)

    def test_ops_daily_actions_skip_stale_same_month_reload(self) -> None:
        for branch_code in WORKFLOW_IDS:
            with self.subTest(branch=branch_code):
                actions = workflow_root_actions(branch_code)
                self.assertEqual(actions["Delete_Existing_Branch_Ops_Daily"]["foreach"], EXPECTED_DELETE_OPS_FOREACH_EXPR)
                self.assertEqual(actions["Apply_to_each_USD_Ops_Daily_Row"]["foreach"], EXPECTED_APPLY_USD_FOREACH_EXPR)
                self.assertEqual(actions["Apply_to_each_CAD_Ops_Daily_Row"]["foreach"], EXPECTED_APPLY_CAD_FOREACH_EXPR)
                analytics_item = actions["Condition_Current_Month_Budget_Record_For_Analytics_Exists"]["actions"][
                    "Update_Current_Month_Budget_Analytics_Payload"
                ]["inputs"]["parameters"]["item"]
                self.assertEqual(analytics_item["qfu_opsdailycadjson"], EXPECTED_PRESERVED_CAD_OPS_EXPR)
                self.assertEqual(analytics_item["qfu_opsdailyusdjson"], EXPECTED_PRESERVED_USD_OPS_EXPR)


if __name__ == "__main__":
    unittest.main()
